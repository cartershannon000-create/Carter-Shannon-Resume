#!/usr/bin/env python3
from __future__ import annotations

import csv
import html
import json
import math
import re
import sqlite3
from calendar import monthrange
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from datetime import date, datetime, timedelta
from pathlib import Path
from statistics import mean
from typing import Any

import openpyxl

try:
    import plaid_sync
except ImportError:  # the dashboard builds fine without the Plaid integration
    plaid_sync = None  # type: ignore[assignment]


ROOT = Path(__file__).resolve().parent
WORKBOOK = ROOT / "Budget_2025-26.xlsx"
SOURCE_DIR = ROOT / "Source data"
OUTPUT = ROOT / "financial_dashboard.html"
DB_PATH = ROOT / "financials.sqlite"
CUTOFF_FOR_CSV_IMPORTS = date(2026, 3, 31)
START_DATE = date(2025, 1, 1)

# --- Plaid (live feed) ----------------------------------------------------------
# Plaid rows come from plaid_store.sqlite, written by plaid_sync.py. They are read
# here exactly like a CSV: a source of Transaction rows, categorised by the same
# rules engine, deduped by the same fingerprint.
#
# Two guards, because a live feed can silently double-count in a way a CSV cannot:
#
# 1. Only `production` data reaches the dashboard. Sandbox is fake money from
#    "First Platypus Bank" and must never mix into real numbers. Sync and preview
#    it all you like; it stops here.
# 2. Plaid and the CSV exports overlap. A charge pulled from Plaid does NOT have
#    the same description as the same charge in a bank CSV, so the dedupe cannot
#    catch it -- it would land twice. PLAID_TAKEOVER_DATE draws a hard line: on and
#    after it, accounts Plaid covers are Plaid's alone and their CSV rows are
#    dropped. Before it, the CSVs stand and Plaid is ignored.
#
# Leaving the date as None means Plaid contributes nothing. That is the safe
# default: an unconfigured live feed changes no number in the dashboard.
PLAID_DASHBOARD_ENV = "production"
# Chosen because the CSV exports stop between 07-25 and 07-31 with no August rows
# at all, so this seam has zero overlap to reconcile. Confirmed independently: the
# derived checking balance at 2026-07-31 (5301.28) plus the single Plaid row after
# it lands exactly on the real bank balance of 7862.56.
PLAID_TAKEOVER_DATE: date | None = date(2026, 8, 1)

# ``mom`` is deliberately NOT internal. It is a contra-expense: money from Mom that
# reimburses spend already booked in a real category (e.g. she covers a flight that is
# sitting in `travel`). Its costs are negative, so leaving it in the ordinary spend pool
# lets it offset those categories and reduce the month's Total, which is how the budget
# workbook has always carried it.
INCOME_CATEGORIES = {"salary"}
TRANSFER_CATEGORIES = {"payment", "venmo", "capone", "amex", "wells"}
SAVINGS_CATEGORIES = {"stock", "additional savings"}
INTERNAL_CATEGORIES = INCOME_CATEGORIES | TRANSFER_CATEGORIES | SAVINGS_CATEGORIES
# Placeholder/junk values that may appear in the workbook's category column. When seen,
# fall back to the rules engine (infer_category) instead of trusting the placeholder.
PLACEHOLDER_CATEGORIES = {"n/a", "b"}

# --- Cash flow (stock and flow) -------------------------------------------------
# Everything above models SPEND, booked when a charge happens. That is an accrual view
# and it deliberately ignores which account the money left and when. The cash flow model
# below answers the other question: what did the checking balance actually do.
#
# The transaction history is pure flow -- no balance is recorded anywhere -- so the stock
# has to be anchored to one known balance and derived outward in both directions.
CASH_ACCOUNT = "Checking"
CARD_ACCOUNTS = {"Amex", "Capital One", "Wells Fargo"}
# Statement balance confirmed by hand. Every other balance in the dashboard is derived
# from this one point by walking the transaction deltas forward and backward.
#
# Validated end to end: walking this anchor forward lands on 5301.28 at 2026-07-31, which
# matched the real account. That matters because the Checking exports have a 10-day hole --
# they run to 05/29 and resume 06/09 -- and the derived balance would be off by whatever
# moved in it. It isn't, so that window is empty. Do not "fix" it by re-pulling May/June.
#
# Note the P&L-to-cash bridge canNOT catch a hole like that: both sides are computed from
# the same loaded rows, so it is an identity that closes even when data is missing. It
# checks the classification, not the coverage. Only a real balance can check coverage.
OPENING_BALANCE_ANCHOR = {"account": CASH_ACCOUNT, "as_of": date(2026, 3, 31), "balance": 5842.31}

CATEGORY_LABELS = {
    "alc": "Alcohol",
    "bar": "Bars",
    "bet": "Betting",
    "cash": "Cash",
    "check": "Checks",
    "clothing": "Clothing",
    "donations": "Donations",
    "fee": "Fees",
    "food": "Food",
    "gift": "Gifts",
    "event": "Events",
    "golf": "Golf",
    "grocery": "Groceries",
    "haircut": "Haircuts",
    "living": "Living",
    "metro": "Metro",
    "mom": "Mom",
    "payment": "Card payments",
    "rent": "Rent",
    "salary": "Salary",
    "service": "Services",
    "spotify": "Spotify",
    "stock": "Investments",
    "additional savings": "Additional Savings",
    "tax": "Taxes",
    "travel": "Travel",
    "uber": "Uber",
    "utilities": "Utilities",
    "venmo": "Venmo transfers",
}


@dataclass
class Transaction:
    id: str
    account: str
    date: str
    month: str
    description: str
    category: str
    category_label: str
    amount: float
    cost: float
    type: str
    status: str
    source: str
    native_category: str = ""
    counterparty: str = ""
    needs_review: int = 0
    # 1-based index of this row among identical rows *within a single source file*.
    # Two identical rows in one export are two real charges; the same row appearing in
    # two overlapping exports is one charge. Both get occurrence 1 in their own file, so
    # they still collapse, while a genuine same-day repeat gets 2 and survives.
    occurrence: int = 1


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def norm_key(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value).upper())


def money(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return 0.0
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    text = text.replace("$", "").replace(",", "").replace("+", "").replace("(", "").replace(")", "").replace(" ", "")
    try:
        parsed = float(text)
    except ValueError:
        return 0.0
    return -parsed if negative else parsed


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt).date()
        except ValueError:
            continue
    return None


def month_key(d: date) -> str:
    return f"{d.year:04d}-{d.month:02d}"


def category_label(category: str) -> str:
    category = (category or "uncategorized").strip().lower()
    return CATEGORY_LABELS.get(category, category.replace("_", " ").title())


def canonical_category(value: Any) -> str:
    text = clean_text(value).lower()
    if not text:
        return "uncategorized"
    aliases = {
        "restaurants": "food",
        "restaurant-bar & cafe": "food",
        "restaurant-bar & café": "food",
        "merchandise & supplies-groceries": "grocery",
        "merchandise & supplies-internet purchase": "living",
        "gas/automotive": "travel",
        "airfare": "travel",
        "travel/entertainment": "travel",
        "other services": "service",
        "fees & adjustments": "fee",
        "internet": "service",
        "gifts": "gift",
        "charity": "donations",
        "check": "cash",
    }
    return aliases.get(text, text)


def build_mapping(wb: openpyxl.Workbook) -> dict[str, str]:
    mapping: dict[str, str] = {}
    ws = wb["Mapping"]
    for row in ws.iter_rows(values_only=True):
        desc = row[1] if len(row) > 1 else None
        cat = row[2] if len(row) > 2 else None
        if desc and cat and clean_text(desc).lower() != "description":
            mapping[norm_key(desc)] = canonical_category(cat)
    return mapping


# Ordered transaction-categorization rules. The FIRST rule whose pattern matches the
# lowercased, normalized transaction description wins, so specific rules come before
# generic ones.
#   "any" -> matches if ANY of the substrings is present
#   "all" -> matches only if ALL of the substrings are present
# The monthly review workflow (skill: chs.monthly-finance-review) appends new merchant
# patterns here whenever a transaction shows up as `uncategorized`. One rule per line.
CATEGORY_RULES: list[dict[str, Any]] = [
    {"cat": "payment", "any": ["online payment", "payment thank you", "automatic payment", "capital one mobile pmt", "american express ach pmt", "credit card auto pay"]},
    {"cat": "venmo", "any": ["venmo cashout", "standard transfer"]},
    {"cat": "mom", "any": ["online transfer from shannon m"]},
    {"cat": "payment", "all": ["online transfer ref", "wells fargo active cash"]},
    # Deloitte fundraiser proceeds are donations, not the Deloitte payroll (which keeps
    # the bare "deloitte" term below). This specific rule must come BEFORE salary.
    {"cat": "donations", "all": ["deloitte", "fundraising"]},
    {"cat": "salary", "any": ["deloitte", "payrll", "payroll"]},
    {"cat": "stock", "any": ["brokerage", "investment auto"]},
    # Rent = only the real landlord Zelle payments (payee was Gazin Eric, then Phil S from
    # Mar 2026). Deliberately NOT a bare "rent" substring, so car rentals do not match here.
    {"cat": "rent", "any": ["zelle to phil s", "zelle to gazin eric"]},
    {"cat": "utilities", "any": ["comcast", "xfinity", "washington gas", "electric", "water gas", "water bill", "internet", "utility"]},
    {"cat": "spotify", "any": ["spotify"]},
    {"cat": "service", "any": ["audible", "apple.com/bill", "amazon prime", "supabase", "openai", "anthropic", "claude", "cloudflare", "wsj", "fly.io", "elevenlabs", "linkedin", "google", "github", "hulu"]},
    {"cat": "clothing", "any": ["the marina at"]},
    {"cat": "living", "any": ["amazon", "amz*", "walmart", "target", "walgreens", "cvs", "home depot", "thd-", "finish line", "wrangler", "tecovas", "peter millar", "connecticut awashington"]},
    {"cat": "grocery", "any": ["harris teeter", "safeway", "trader joe", "whole foods", "wholefds", "costco", "grocery", "groceries", "kroger", "market", "kalorama mark", "connecticut avenue", "lucky seven", "dupont circle", "clark's marke"]},
    # Bars, ahead of food. "tst*" (Toast POS) sits in the food rule and is used by bars
    # and restaurants alike, so without this it silently claimed every Toast bar --
    # including ones whose own token is already in the bar rule below (Hank's Oyster Bar
    # matched "tst*" before it could match "hanks oy"). Policy: if a venue is a bar, its
    # spend is bar spend. Only high-confidence tokens belong here; a bare "bar" would hit
    # "rhubarb"/"barnes", and "lounge" would steal "capital 1 lounge" from travel.
    {"cat": "bar", "any": [
        "beer", "brewing", "brewery", "brewpub", "taproom", "tap room", "tavern",
        "biergarten", "beer garden", "saloon", "distillery", "cocktail", "speakeasy",
        "oyster bar", "wine bar", "sports bar", "pub ", " pub", "alehouse", "ale house",
        # Known local bars whose names carry no generic bar token.
        "tune inn", "hanks oy", "aslin", "dacha", "exiles", "the eldo", "wet dog",
        "universal win", "blaguard", "vegas lounge", "kardinal", "doyle bar",
        "players club", "easy tiger",
        # Confirmed from the budget workbook, where these were hand-categorised as bar.
        # "garden d" alone is avoided -- it would also catch "OLIVE GARDEN DC".
        "commodore", "garden district", "garden dwashington",
        # Confirmed by hand. "salt line" and "whiskey jar" were previously in the food
        # rule below and had to move up here; Cana went the other way, to food.
        "casamara", "salt line", "whiskey jar",
    ]},
    {"cat": "food", "any": ["uber eats", "doordash", "domino", "chipotle", "cava", "restaurant", "cafe", "coffee", "kramerbooks", "tst*", "bagel", "panera", "shake shack", "starbucks", "cana dc", "dukes", "taco bell", "raising canes", "doner", "wok and roll", "teocalli", "bibibop", "levy@", "johnny?s gour", "lauriol", "food", "burger", "mcdonald", "dupont eatery", "mozzo", "larrys cookie", "van leeuwen", "🍔", "☕"]},
    {"cat": "bar", "any": ["beer", "bar", "wine", "liquor", "hanks oy", "universal win", "wet dog", "the eldo", "aplpay store washington", "drinks", "drink", "🍻", "🍺", "🍷", "🥂"]},
    {"cat": "uber", "any": ["uber", "lyft", "uvc inc"]},
    {"cat": "metro", "any": ["metro", "wmata"]},
    {"cat": "bet", "any": ["draftking", "fanduel", "kalshi", "bet"]},
    {"cat": "golf", "any": ["golf", "titleist", "glf*"]},
    # Travel includes car rentals ("rent a car" etc.) so they do not fall into rent.
    # " inn" has a leading space so it matches "Old Town Inn" but not "dinner"/"winner".
    {"cat": "travel", "any": ["airline", "american airl", "hotel", " inn", "airbnb", "flight", "airfare", "travel", "circle k", "rent a car", "car rent", "rental", "budget rent", "capital 1 lounge", "american express tra"]},
    # event: two original rules merged (same category, were adjacent).
    {"cat": "event", "any": ["bowling", "bowl", "pinstripes", "five iron", "5i", "railbird", "karaoke", "strip club", "the atlant", "photo-ma", "ufc white", "650 industries", "tm - wdc host", "muzette karaoke", "bball", "defending"]},
    {"cat": "utilities", "any": ["cleaning", "hand soap", "soap", "trash bags", "paper towels", "tinfoil"]},
    # fee: two original rules merged (same category, were adjacent).
    {"cat": "fee", "any": ["sec of state", "national link", "wire fee", "red light"]},
    {"cat": "donations", "any": ["washington and lee"]},
    {"cat": "haircut", "any": ["nick zhang", "💅"]},
    {"cat": "living", "any": ["air fryer"]},
    {"cat": "gift", "any": ["gift", "book for your dad", "sp twigs", "sp manse", "🎁"]},
    {"cat": "donations", "any": ["appeal", "donation", "charity", "inspire inc", "paypal transfer"]},
    {"cat": "cash", "any": ["atm withdrawal"]},
]


def infer_category(description: str, native: str = "", mapping: dict[str, str] | None = None) -> str:
    key = norm_key(description)
    if mapping and key in mapping:
        return mapping[key]

    text = key.lower()
    for rule in CATEGORY_RULES:
        if "all" in rule:
            if all(term in text for term in rule["all"]):
                return rule["cat"]
        elif any(term in text for term in rule["any"]):
            return rule["cat"]

    native_key = canonical_category(native)
    if native_key and native_key != "uncategorized":
        return native_key
    return "uncategorized"


# Description-pattern overrides that take precedence over the source-provided category.
# The budget workbook lumped some non-salary inflows under "salary": IRS/state tax
# refunds and payments (now their own 'tax' category), and a couple of deposited checks
# (treated as cash/reimbursements). These correct that at import time.
OVERRIDE_RULES: list[dict[str, Any]] = [
    {"cat": "tax", "any": ["usataxpymt", "tax ref", "taxpayment", "taxrfd", "mef-iit", "dcref-iit"]},
]


# Venues where food-vs-bar is a genuine coin flip: restaurants with a serious bar program,
# or Toast (tst*) merchants whose truncated name gives no signal either way. These still
# get a best-guess category, but are flagged so they surface in the Review tab to be
# confirmed by hand rather than silently landing in the wrong bucket.
#
# Currently empty: every merchant that was in here has since been settled by hand and
# given a real rule above (Casamara/Salt Line/Whiskey Jar to bar, Cana to food). The
# machinery stays so the next batch of unclear merchants can be routed to Review rather
# than silently guessed at -- add substrings here and they show up there.
AMBIGUOUS_FOOD_BAR: list[str] = []


def is_ambiguous_food_bar(description: str, category: str) -> bool:
    """True when a food/bar call is a toss-up and should be confirmed by hand."""
    if category not in {"food", "bar"}:
        return False
    text = norm_key(description).lower()
    return any(term in text for term in AMBIGUOUS_FOOD_BAR)


def override_category(description: str, category: str) -> str:
    text = norm_key(description).lower()
    for rule in OVERRIDE_RULES:
        if any(term in text for term in rule["any"]):
            return rule["cat"]
    if "mobile deposit" in text and category == "salary":
        return "cash"  # a deposited check the workbook mislabeled as salary
    return category


def add_tx(
    transactions: list[Transaction],
    *,
    account: str,
    d: date | None,
    description: Any,
    category: Any,
    amount: Any,
    cost: Any,
    tx_type: str = "",
    status: str = "",
    source: str,
    native_category: str = "",
    counterparty: str = "",
    occurrence: int = 1,
) -> None:
    if d is None or d < START_DATE:
        return
    desc = clean_text(description)
    if not desc:
        desc = f"{account} {clean_text(category) or 'transaction'}"
    cat = canonical_category(category)
    if cat in PLACEHOLDER_CATEGORIES:
        cat = infer_category(desc, native_category)
    cat = override_category(desc, cat)
    if account == "Amex" and cat == "mom" and any(term in desc.lower() for term in ("hotel", "amextravel", "travel")):
        cat = "travel"
    if cat == "uncategorized" and account == "Venmo":
        # Person-to-person Venmo defaults to a food expense, unless it is to Nick (haircut).
        cat = "haircut" if "nick" in clean_text(counterparty).lower() else "food"
    amt = round(money(amount), 2)
    cst = round(money(cost), 2)
    party = clean_text(counterparty)
    tx_id = f"{account}|{d.isoformat()}|{desc}|{amt:.2f}|{source}"
    # Four roommates sending the same amount with the same note on the same day differ
    # only by counterparty, so it has to be part of the id or they collide on the primary
    # key. Only appended when present (Venmo), leaving every other account's ids
    # unchanged. Occurrence is suffixed only from the 2nd repeat on, for the same reason.
    if party:
        tx_id = f"{tx_id}|{party}"
    if occurrence > 1:
        tx_id = f"{tx_id}|#{occurrence}"
    transactions.append(
        Transaction(
            id=tx_id,
            account=account,
            date=d.isoformat(),
            month=month_key(d),
            description=desc,
            category=cat,
            category_label=category_label(cat),
            amount=amt,
            cost=cst,
            type=clean_text(tx_type),
            status=clean_text(status),
            source=source,
            native_category=clean_text(native_category),
            counterparty=party,
            needs_review=1 if is_ambiguous_food_bar(desc, cat) else 0,
            occurrence=occurrence,
        )
    )


def parse_workbook_transactions(wb: openpyxl.Workbook) -> list[Transaction]:
    transactions: list[Transaction] = []

    for sheet, account, start_row in [
        ("Checking", "Checking", 4),
        ("Wells", "Wells Fargo", 3),
    ]:
        ws = wb[sheet]
        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            add_tx(
                transactions,
                account=account,
                d=parse_date(row[2] if len(row) > 2 else None),
                amount=row[3] if len(row) > 3 else None,
                description=row[6] if len(row) > 6 else None,
                category=row[14] if len(row) > 14 else None,
                cost=row[15] if len(row) > 15 else None,
                source=f"Workbook:{sheet}:{row_idx}",
            )

    ws = wb["CapOne"]
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        add_tx(
            transactions,
            account="Capital One",
            d=parse_date(row[2] if len(row) > 2 else None),
            amount=row[12] if len(row) > 12 else row[7] if len(row) > 7 else None,
            description=row[5] if len(row) > 5 else None,
            category=row[14] if len(row) > 14 else None,
            cost=row[15] if len(row) > 15 else None,
            source=f"Workbook:CapOne:{row_idx}",
            native_category=clean_text(row[6] if len(row) > 6 else ""),
        )

    ws = wb["Amex"]
    for row_idx, row in enumerate(ws.iter_rows(min_row=6, values_only=True), start=6):
        add_tx(
            transactions,
            account="Amex",
            d=parse_date(row[2] if len(row) > 2 else None),
            amount=row[4] if len(row) > 4 else None,
            description=row[3] if len(row) > 3 else None,
            category=row[14] if len(row) > 14 else None,
            cost=row[15] if len(row) > 15 else None,
            source=f"Workbook:Amex:{row_idx}",
            native_category=clean_text(row[12] if len(row) > 12 else ""),
        )

    ws = wb["Venmo"]
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        p1 = clean_text(row[7] if len(row) > 7 else "")
        p2 = clean_text(row[8] if len(row) > 8 else "")
        add_tx(
            transactions,
            account="Venmo",
            d=parse_date(row[2] if len(row) > 2 else None),
            amount=row[9] if len(row) > 9 else None,
            description=row[6] if len(row) > 6 else None,
            category=row[14] if len(row) > 14 else None,
            cost=row[15] if len(row) > 15 else None,
            tx_type=row[4] if len(row) > 4 else "",
            status=row[5] if len(row) > 5 else "",
            source=f"Workbook:Venmo:{row_idx}",
            counterparty=" / ".join(part for part in (p1, p2) if part),
        )

    return transactions


def parse_recent_csvs(mapping: dict[str, str]) -> list[Transaction]:
    transactions: list[Transaction] = []

    def after_cutoff(d: date | None) -> bool:
        return bool(d and d > CUTOFF_FOR_CSV_IMPORTS)

    # Read current exports plus files moved into archive subfolders. This preserves
    # earlier post-workbook history when a bank export is replaced with a newer file.
    def csv_paths(pattern: str) -> list[Path]:
        return sorted(SOURCE_DIR.rglob(pattern))

    for path in csv_paths("Checking.csv"):
        seen_in_file: Counter[tuple] = Counter()
        with path.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                d = parse_date(row.get("DATE"))
                if not after_cutoff(d):
                    continue
                desc = row.get("DESCRIPTION", "")
                amount = money(row.get("AMOUNT"))
                cat = infer_category(desc, mapping=mapping)
                seen_in_file[(d, clean_text(desc), round(amount, 2))] += 1
                add_tx(
                    transactions,
                    account="Checking",
                    d=d,
                    description=desc,
                    category=cat,
                    amount=amount,
                    cost=-amount,
                    status=row.get("STATUS", ""),
                    source="CSV:Checking",
                    occurrence=seen_in_file[(d, clean_text(desc), round(amount, 2))],
                )

    for path in csv_paths("CreditCard.csv"):
        seen_in_file = Counter()
        with path.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                d = parse_date(row.get("DATE"))
                if not after_cutoff(d):
                    continue
                desc = row.get("DESCRIPTION", "")
                amount = money(row.get("AMOUNT"))
                cat = infer_category(desc, mapping=mapping)
                seen_in_file[(d, clean_text(desc), round(amount, 2))] += 1
                add_tx(
                    transactions,
                    account="Wells Fargo",
                    d=d,
                    description=desc,
                    category=cat,
                    amount=amount,
                    cost=-amount,
                    status=row.get("STATUS", ""),
                    source="CSV:CreditCard",
                    occurrence=seen_in_file[(d, clean_text(desc), round(amount, 2))],
                )

    for path in csv_paths("*_transaction_download.csv"):
        seen_in_file = Counter()
        with path.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                d = parse_date(row.get("Transaction Date"))
                if not after_cutoff(d):
                    continue
                desc = row.get("Description", "")
                debit = money(row.get("Debit"))
                credit = money(row.get("Credit"))
                amount = debit - credit
                cat = infer_category(desc, row.get("Category", ""), mapping)
                seen_in_file[(d, clean_text(desc), round(amount, 2))] += 1
                add_tx(
                    transactions,
                    account="Capital One",
                    d=d,
                    description=desc,
                    category=cat,
                    amount=amount,
                    cost=amount,
                    source="CSV:CapitalOne",
                    native_category=row.get("Category", ""),
                    occurrence=seen_in_file[(d, clean_text(desc), round(amount, 2))],
                )

    for path in csv_paths("activity.csv"):
        seen_in_file = Counter()
        with path.open(newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                d = parse_date(row.get("Date"))
                if not after_cutoff(d):
                    continue
                desc = row.get("Description", "")
                amount = money(row.get("Amount"))
                cat = infer_category(desc, mapping=mapping)
                seen_in_file[(d, clean_text(desc), round(amount, 2))] += 1
                add_tx(
                    transactions,
                    account="Amex",
                    d=d,
                    description=desc,
                    category=cat,
                    amount=amount,
                    cost=amount,
                    source="CSV:AmexActivity",
                    occurrence=seen_in_file[(d, clean_text(desc), round(amount, 2))],
                )

    for path in csv_paths("VenmoStatement_*.csv"):
        seen_in_file = Counter()
        with path.open(newline="", encoding="utf-8-sig") as handle:
            rows = list(csv.reader(handle))
        if len(rows) < 4:
            continue
        header = rows[2]
        columns = {name: idx for idx, name in enumerate(header)}
        for raw in rows[4:]:
            if len(raw) <= columns.get("Datetime", 99):
                continue
            d = parse_date(raw[columns["Datetime"]])
            if not after_cutoff(d):
                continue
            desc = raw[columns.get("Note", 0)] if columns.get("Note") is not None else ""
            amount = money(raw[columns.get("Amount (total)", 0)] if columns.get("Amount (total)") is not None else "")
            from_person = raw[columns.get("From", 0)] if columns.get("From") is not None else ""
            to_person = raw[columns.get("To", 0)] if columns.get("To") is not None else ""
            tx_type = raw[columns.get("Type", 0)] if columns.get("Type") is not None else ""
            cat = infer_category(" ".join([desc, tx_type, from_person, to_person]), mapping=mapping)
            party = " / ".join(part for part in (from_person, to_person) if part)
            seen_in_file[(d, clean_text(desc), round(amount, 2), clean_text(party))] += 1
            add_tx(
                transactions,
                account="Venmo",
                d=d,
                description=desc or tx_type,
                category=cat,
                amount=amount,
                cost=-amount,
                tx_type=tx_type,
                status=raw[columns.get("Status", 0)] if columns.get("Status") is not None else "",
                source=f"CSV:{path.name}",
                counterparty=party,
                occurrence=seen_in_file[(d, clean_text(desc), round(amount, 2), clean_text(party))],
            )

    return transactions


# Plaid's personal_finance_category taxonomy translated into this dashboard's
# vocabulary. This table is not optional decoration: infer_category() falls back
# to whatever native category a source supplies, which is safe for Capital One's
# dozen curated values but not for Plaid's 104. Left untranslated, a single
# month invents categories like "General Merchandise Gifts And Novelties",
# which the budget comparison and the Review tab have no idea what to do with.
#
# Anything absent from both tables stays `uncategorized` and surfaces in Review.
# That is the point: a wrong guess is worse than an explicit question.
PLAID_CATEGORY_BY_DETAILED = {
    "INCOME_WAGES": "salary",
    "INCOME_TAX_REFUND": "tax",
    "TRANSFER_OUT_INVESTMENT_AND_RETIREMENT_FUNDS": "stock",
    "TRANSFER_OUT_SAVINGS": "additional savings",
    "TRANSFER_OUT_WITHDRAWAL": "cash",
    "LOAN_PAYMENTS_CREDIT_CARD_PAYMENT": "payment",
    "ENTERTAINMENT_CASINOS_AND_GAMBLING": "bet",
    "ENTERTAINMENT_SPORTING_EVENTS_AMUSEMENT_PARKS_AND_MUSEUMS": "event",
    "FOOD_AND_DRINK_BEER_WINE_AND_LIQUOR": "alc",
    "FOOD_AND_DRINK_COFFEE": "food",
    "FOOD_AND_DRINK_FAST_FOOD": "food",
    "FOOD_AND_DRINK_GROCERIES": "grocery",
    "FOOD_AND_DRINK_RESTAURANT": "food",
    "FOOD_AND_DRINK_VENDING_MACHINES": "food",
    "GENERAL_MERCHANDISE_CLOTHING_AND_ACCESSORIES": "clothing",
    "GENERAL_MERCHANDISE_GIFTS_AND_NOVELTIES": "gift",
    "PERSONAL_CARE_HAIR_AND_BEAUTY": "haircut",
    "GOVERNMENT_AND_NON_PROFIT_DONATIONS": "donations",
    "GOVERNMENT_AND_NON_PROFIT_TAX_PAYMENT": "tax",
    "TRANSPORTATION_TAXIS_AND_RIDE_SHARES": "uber",
    "TRANSPORTATION_PUBLIC_TRANSIT": "metro",
    # Lime/Bikeshare are point-to-point paid rides, so they ride along with Uber
    # rather than becoming a category of their own. Flip this line if they should
    # sit with Metro instead.
    "TRANSPORTATION_BIKES_AND_SCOOTERS": "uber",
    "TRANSPORTATION_GAS": "travel",
    "TRANSPORTATION_PARKING": "travel",
    "TRANSPORTATION_TOLLS": "travel",
    "RENT_AND_UTILITIES_RENT": "rent",
    "RENT_AND_UTILITIES_INTERNET_AND_CABLE": "service",
    "RENT_AND_UTILITIES_TELEPHONE": "service",
}

# Used only when the detailed value is unmapped. Restricted on purpose to the
# primaries where a miss is merely a mis-bucketed expense. INCOME, TRANSFER_IN,
# TRANSFER_OUT and LOAN_PAYMENTS are deliberately excluded -- those decide
# income, savings and card-payment accounting, and guessing there would quietly
# distort the cash flow bridge rather than just a spend line.
PLAID_CATEGORY_BY_PRIMARY = {
    "BANK_FEES": "fee",
    "FOOD_AND_DRINK": "food",
    "GENERAL_MERCHANDISE": "living",
    "GENERAL_SERVICES": "service",
    "HOME_IMPROVEMENT": "living",
    "PERSONAL_CARE": "service",
    "RENT_AND_UTILITIES": "utilities",
    "TRAVEL": "travel",
}


def plaid_category(detailed: str, primary: str) -> str:
    """Translate a Plaid category, or return 'uncategorized' to force a review."""
    detailed = (detailed or "").strip().upper()
    primary = (primary or "").strip().upper()
    if detailed in PLAID_CATEGORY_BY_DETAILED:
        return PLAID_CATEGORY_BY_DETAILED[detailed]
    return PLAID_CATEGORY_BY_PRIMARY.get(primary, "uncategorized")


def parse_plaid_transactions(mapping: dict[str, str]) -> list[Transaction]:
    """Read the Plaid store and turn it into Transaction rows.

    Returns nothing at all unless PLAID_TAKEOVER_DATE is set -- see the guards at
    the top of this file.
    """
    if plaid_sync is None or PLAID_TAKEOVER_DATE is None:
        return []

    rows = plaid_sync.load_dashboard_rows(env=PLAID_DASHBOARD_ENV)
    transactions: list[Transaction] = []
    seen_in_feed: Counter[tuple] = Counter()

    # Sorted by transaction id so the occurrence numbers a repeat charge gets are
    # stable between builds; otherwise the same charge changes primary key when
    # the rows come back in a different order and the dashboard shows a new row.
    for row in sorted(rows, key=lambda r: (r["date"], r["transaction_id"])):
        d = parse_date(row["date"])
        if d is None or d < PLAID_TAKEOVER_DATE:
            continue
        desc = row["description"]
        # Description rules first and alone -- they are tuned to the merchants
        # that actually appear on these statements and beat any generic
        # taxonomy. The native category is withheld here so infer_category
        # cannot fall back to a raw Plaid enum; that fallback is replaced by the
        # translation table below.
        cat = infer_category(desc, "", mapping)
        if cat == "uncategorized":
            cat = plaid_category(row["category_detailed"], row["category_primary"])
        key = (d, clean_text(desc), round(row["amount"], 2))
        seen_in_feed[key] += 1
        add_tx(
            transactions,
            account=row["account"],
            d=d,
            description=desc,
            category=cat,
            amount=row["amount"],
            cost=row["cost"],
            source=f"Plaid:{row['account']}",
            native_category=row["native_category"],
            counterparty=row["merchant_name"],
            occurrence=seen_in_feed[key],
        )
    return transactions


def plaid_covered_accounts() -> set[str]:
    """Dashboard accounts that Plaid is currently feeding."""
    if plaid_sync is None or PLAID_TAKEOVER_DATE is None:
        return set()
    return {
        b["dashboard_account"]
        for b in plaid_sync.load_balances(env=PLAID_DASHBOARD_ENV)
        if b["dashboard_account"]
    }


def apply_plaid_takeover(transactions: list[Transaction]) -> tuple[list[Transaction], int]:
    """Drop CSV/workbook rows for accounts and dates that Plaid now owns.

    Without this the same charge arrives twice with two different descriptions --
    the bank's export wording and Plaid's -- which the dedupe fingerprint cannot
    see as the same thing.
    """
    covered = plaid_covered_accounts()
    if not covered or PLAID_TAKEOVER_DATE is None:
        return transactions, 0
    cutoff = PLAID_TAKEOVER_DATE.isoformat()
    kept = [
        tx for tx in transactions
        if tx.source.startswith("Plaid:") or tx.account not in covered or tx.date < cutoff
    ]
    return kept, len(transactions) - len(kept)


def tx_is_expense(tx: Transaction) -> bool:
    return tx.category not in INTERNAL_CATEGORIES


def spend_cost(tx: Transaction) -> float:
    return tx.cost if tx_is_expense(tx) else 0.0


def income_amount(tx: Transaction) -> float:
    return max(-tx.cost, 0.0) if tx.category in INCOME_CATEGORIES else 0.0


def savings_amount(tx: Transaction) -> float:
    return max(tx.cost, 0.0) if tx.category in SAVINGS_CATEGORIES else 0.0


def build_summary(transactions: list[Transaction]) -> dict[str, Any]:
    transactions = sorted(transactions, key=lambda tx: (tx.date, tx.account, tx.description))
    months = sorted({tx.month for tx in transactions})
    latest_month = months[-1] if months else ""
    complete_months = [m for m in months if m < latest_month] or months
    recent_basis = complete_months[-3:] if len(complete_months) >= 3 else complete_months

    by_month: dict[str, dict[str, Any]] = {}
    for month in months:
        rows = [tx for tx in transactions if tx.month == month]
        category_spend = defaultdict(float)
        for tx in rows:
            if tx_is_expense(tx):
                category_spend[tx.category] += tx.cost
        top_category = max(category_spend.items(), key=lambda item: item[1], default=("", 0.0))
        expenses = sum(spend_cost(tx) for tx in rows)
        income = sum(income_amount(tx) for tx in rows)
        savings = sum(savings_amount(tx) for tx in rows)
        by_month[month] = {
            "month": month,
            "income": round(income, 2),
            "expenses": round(expenses, 2),
            "rent": round(sum(tx.cost for tx in rows if tx.category == "rent"), 2),
            "savings": round(savings, 2),
            "net_after_spend_and_savings": round(income - expenses - savings, 2),
            "transactions": len(rows),
            "top_category": category_label(top_category[0]) if top_category[0] else "",
            "top_category_amount": round(top_category[1], 2),
        }

    by_account = []
    for account in sorted({tx.account for tx in transactions}):
        rows = [tx for tx in transactions if tx.account == account]
        by_account.append(
            {
                "account": account,
                "count": len(rows),
                "first": min(tx.date for tx in rows),
                "last": max(tx.date for tx in rows),
                "expense": round(sum(spend_cost(tx) for tx in rows), 2),
                "income": round(sum(income_amount(tx) for tx in rows), 2),
                "savings": round(sum(savings_amount(tx) for tx in rows), 2),
            }
        )

    by_category = []
    for category in sorted({tx.category for tx in transactions}):
        rows = [tx for tx in transactions if tx.category == category]
        expense = round(sum(spend_cost(tx) for tx in rows), 2)
        income = round(sum(income_amount(tx) for tx in rows), 2)
        savings = round(sum(savings_amount(tx) for tx in rows), 2)
        by_category.append(
            {
                "category": category,
                "label": category_label(category),
                "count": len(rows),
                "expense": expense,
                "income": income,
                "savings": savings,
            }
        )
    by_category.sort(key=lambda row: row["expense"], reverse=True)

    overall = {
        "first_date": min((tx.date for tx in transactions), default=""),
        "last_date": max((tx.date for tx in transactions), default=""),
        "transaction_count": len(transactions),
        "income": round(sum(income_amount(tx) for tx in transactions), 2),
        "expenses": round(sum(spend_cost(tx) for tx in transactions), 2),
        "savings": round(sum(savings_amount(tx) for tx in transactions), 2),
        "rent": round(sum(tx.cost for tx in transactions if tx.category == "rent"), 2),
        "account_count": len({tx.account for tx in transactions}),
    }
    overall["net_after_spend_and_savings"] = round(overall["income"] - overall["expenses"] - overall["savings"], 2)
    overall["monthly_average_expense"] = round(mean([by_month[m]["expenses"] for m in complete_months]), 2) if complete_months else 0
    overall["recent_monthly_average_expense"] = round(mean([by_month[m]["expenses"] for m in recent_basis]), 2) if recent_basis else 0

    return {
        "overall": overall,
        "months": [by_month[m] for m in months],
        "accounts": by_account,
        "categories": by_category,
        "recent_basis": recent_basis,
    }


def _percentile(values: list[float], fraction: float) -> float:
    """Match PostgreSQL percentile_cont so the Python and SQL forecasts stay identical."""
    ordered = sorted(values)
    if not ordered:
        return 0.0
    position = (len(ordered) - 1) * fraction
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    return ordered[lower] + (ordered[upper] - ordered[lower]) * (position - lower)


def normalize_merchant(description: str) -> str:
    """Collapse volatile transaction descriptions to the merchant identity Insights uses."""
    merchant = re.sub(r"[^a-z0-9 ]", "", description.lower())
    merchant = re.sub(r"\b[0-9a-z]{5,}\b", "", merchant)
    return " ".join(merchant.split()[:3])


def forecast_merchant_label(tx: Transaction) -> tuple[int, str]:
    """Return a display label without changing the recurring-charge grouping key."""
    counterparty = clean_text(tx.counterparty)
    if counterparty:
        return (1, counterparty)

    label = clean_text(tx.description)
    label = re.sub(
        r"^zelle\s+|\s+on\s+\d{1,2}/\d{1,2}\s+ref\b.*$",
        "",
        label,
        flags=re.IGNORECASE,
    )
    label = re.sub(r"\b\d{3}[- ]\d{3}[- ]\d{4}\b", " ", label)
    # Remove reference-like alphanumeric tokens and terminal digit runs, but retain the
    # ordinary words that the grouping normaliser intentionally discards (for example,
    # SPOTIFY). Counterparty labels win over all description fallbacks for a merchant.
    label = re.sub(r"\b(?=[0-9a-z]*\d)[0-9a-z]{5,}\b", " ", label, flags=re.IGNORECASE)
    label = re.sub(r"(?:\s+#?\d+)+\s*$", "", label)
    label = clean_text(label).strip(" -–—*#:/")[:28].rstrip().title()
    return (0, label)


FORECAST_MODEL_VERSION = "daytype-1"


def _build_forecast_daytype_base(transactions: list[Transaction]) -> dict[str, Any]:
    """Forecast committed charges and calendar-sensitive variable category spend."""
    today = date.today()
    current_month = month_key(today)
    day_of_month = today.day
    days_in_month = monthrange(today.year, today.month)[1]
    month_end = date(today.year, today.month, days_in_month)
    future_dates = [today + timedelta(days=offset) for offset in range(1, (month_end - today).days + 1)]
    weekdays_remaining = sum(day.weekday() < 5 for day in future_dates)
    weekend_days_remaining = len(future_dates) - weekdays_remaining
    expense_rows = [tx for tx in transactions if tx_is_expense(tx)]
    categories = sorted({tx.category for tx in expense_rows})
    complete_months = sorted({tx.month for tx in expense_rows if tx.month < current_month})
    month_cursor = date(today.year, today.month, 1)
    trailing_months = []
    for _ in range(6):
        month_cursor = (month_cursor - timedelta(days=1)).replace(day=1)
        trailing_months.append(month_key(month_cursor))
    trailing_months.reverse()
    # Keep detection and its residual variable history on one window. Otherwise an old
    # merchant alias (the prior rent recipient) survives removal and forecasts a second
    # copy of the same obligation even though the current alias is already committed.
    forecast_months = trailing_months or complete_months

    labels = {
        tx.category: tx.category_label or category_label(tx.category)
        for tx in expense_rows
    }
    merchant_months: dict[tuple[str, str], dict[str, list[Transaction]]] = defaultdict(lambda: defaultdict(list))
    merchant_label_counts: dict[str, Counter[tuple[int, str]]] = defaultdict(Counter)
    for tx in expense_rows:
        merchant = normalize_merchant(tx.description)
        if tx.month in trailing_months and tx.cost > 0 and merchant:
            merchant_months[(merchant, tx.category)][tx.month].append(tx)
            priority, merchant_label = forecast_merchant_label(tx)
            if merchant_label:
                merchant_label_counts[merchant][(priority, merchant_label)] += 1

    merchant_labels: dict[str, str] = {}
    for merchant, counts in merchant_label_counts.items():
        # Counterparty-derived candidates (priority 1) always beat description-derived
        # candidates. Within that source, use the most frequent value and a stable text
        # tie-break so row order cannot change the label.
        merchant_labels[merchant] = min(
            counts,
            key=lambda candidate: (
                -candidate[0], -counts[candidate], candidate[1].casefold(), candidate[1]
            ),
        )[1]

    committed_definitions: list[dict[str, Any]] = []
    trailing_count = len(trailing_months)
    for (merchant, category), by_month in merchant_months.items():
        months_seen = len(by_month)
        if months_seen < 3 or not trailing_count or months_seen / trailing_count < 0.6:
            continue
        # A merchant may split one monthly obligation across multiple ledger rows (rent
        # does this once). Treat the monthly sum as the observed charge so the small
        # companion row cannot make an otherwise fixed obligation look volatile.
        monthly_amounts = [sum(tx.cost for tx in rows) for rows in by_month.values()]
        expected_amount = _percentile(monthly_amounts, 0.5)
        average_amount = mean(monthly_amounts)
        stddev = math.sqrt(mean([(amount - average_amount) ** 2 for amount in monthly_amounts]))
        identical = max(monthly_amounts) - min(monthly_amounts) < 0.005
        if not identical and (expected_amount <= 0 or stddev / expected_amount > 0.15):
            continue
        monthly_days = [
            _percentile([float(int(tx.date[-2:])) for tx in rows], 0.5)
            for rows in by_month.values()
        ]
        expected_day = min(days_in_month, int(math.floor(_percentile(monthly_days, 0.5) + 0.5)))
        committed_definitions.append({
            "merchant_key": merchant,
            "merchant": merchant.title(),
            "merchant_label": merchant_labels.get(merchant, merchant.title()),
            "category": category,
            "label": labels.get(category, category_label(category)),
            "expected_amount": round(expected_amount, 2),
            "expected_day": expected_day,
            "months_seen": months_seen,
        })

    committed_keys = {
        (item["merchant_key"], item["category"])
        for item in committed_definitions
    }
    committed_merchant_keys = {item["merchant_key"] for item in committed_definitions}
    historical: dict[str, dict[str, list[Transaction]]] = defaultdict(lambda: defaultdict(list))
    current: dict[str, list[Transaction]] = defaultdict(list)
    current_variable: dict[str, list[Transaction]] = defaultdict(list)
    variable_excluded_from_history = 0
    for tx in expense_rows:
        merchant_key = normalize_merchant(tx.description)
        key = (merchant_key, tx.category)
        if tx.month < current_month:
            if merchant_key not in committed_merchant_keys:
                historical[tx.category][tx.month].append(tx)
            elif tx.month in forecast_months:
                variable_excluded_from_history += 1
        elif tx.month == current_month and tx.date <= today.isoformat():
            current[tx.category].append(tx)
            if merchant_key not in committed_merchant_keys:
                current_variable[tx.category].append(tx)

    current_committed: dict[tuple[str, str], list[Transaction]] = defaultdict(list)
    for tx in expense_rows:
        key = (normalize_merchant(tx.description), tx.category)
        if tx.month == current_month and tx.date <= today.isoformat() and key in committed_keys:
            current_committed[key].append(tx)

    committed_items: list[dict[str, Any]] = []
    committed_by_category: dict[str, float] = defaultdict(float)
    committed_charged_by_category: dict[str, float] = defaultdict(float)
    for definition in committed_definitions:
        key = (definition["merchant_key"], definition["category"])
        charged_rows = current_committed.get(key, [])
        if charged_rows:
            status = "charged"
            charged_amount = sum(tx.cost for tx in charged_rows)
            committed_charged_by_category[definition["category"]] += charged_amount
        else:
            status = "overdue" if definition["expected_day"] < day_of_month else "due"
            committed_by_category[definition["category"]] += definition["expected_amount"]
        committed_items.append({
            key: value for key, value in definition.items() if key != "merchant_key"
        } | {"status": status})

    status_order = {"overdue": 0, "due": 1, "charged": 2}
    committed_items.sort(key=lambda item: (status_order[item["status"]], item["expected_day"], item["merchant_label"]))

    category_rows: list[dict[str, Any]] = []
    for category in categories:
        months_of_history = len(forecast_months)
        basis = "profile" if months_of_history >= 3 else "run_rate"
        spent = round(sum(tx.cost for tx in current[category]), 2)
        variable_spent = sum(tx.cost for tx in current_variable[category])
        committed_remaining = committed_by_category[category]

        historical_totals: list[float] = []
        months_with_any_spend = 0
        weekday_values: list[float] = []
        weekend_values: list[float] = []
        for month in forecast_months:
            rows = historical[category].get(month, [])
            historical_totals.append(sum(tx.cost for tx in rows))
            months_with_any_spend += bool(rows)
            year, month_number = (int(part) for part in month.split("-"))
            calendar_days = [date(year, month_number, day) for day in range(1, monthrange(year, month_number)[1] + 1)]
            weekday_count = sum(day.weekday() < 5 for day in calendar_days)
            weekend_count = len(calendar_days) - weekday_count
            weekday_values.append(
                sum(tx.cost for tx in rows if date.fromisoformat(tx.date).weekday() < 5) / weekday_count
            )
            weekend_values.append(
                sum(tx.cost for tx in rows if date.fromisoformat(tx.date).weekday() >= 5) / weekend_count
            )

        # Each observation is a month's day-type total divided by every matching
        # calendar day in that month. Days without transactions therefore remain in the
        # denominator instead of silently inflating sparse categories' rates.
        weekday_rates = {p: _percentile(weekday_values, p) for p in (0.25, 0.5, 0.75)}
        weekend_rates = {p: _percentile(weekend_values, p) for p in (0.25, 0.5, 0.75)}

        # Seasonality needs more history than the six-month rate window. Each prior
        # observation is compared with the trailing-twelve monthly average available at
        # that point; incomplete trailing windows are not treated as twelve-month data.
        category_variable_rows = [
            tx for tx in expense_rows
            if tx.category == category
            and tx.month < current_month
            and normalize_merchant(tx.description) not in committed_merchant_keys
        ]
        seasonal_factor = 1.0
        seasonal_months_observed = 0
        if category_variable_rows:
            first_year, first_month = (int(part) for part in min(tx.month for tx in category_variable_rows).split("-"))
            cursor = date(first_year, first_month, 1)
            last_complete = date(today.year, today.month, 1) - timedelta(days=1)
            monthly_variable: list[tuple[date, float]] = []
            variable_by_month = defaultdict(float)
            for tx in category_variable_rows:
                variable_by_month[tx.month] += tx.cost
            while cursor <= last_complete:
                monthly_variable.append((cursor, variable_by_month[cursor.strftime("%Y-%m")]))
                cursor = date(cursor.year + (cursor.month == 12), cursor.month % 12 + 1, 1)

            history_months = len(monthly_variable)
            seasonal_ratios = []
            for index, (month_start, month_total) in enumerate(monthly_variable):
                if month_start.month != today.month or index < 11:
                    continue
                trailing_average = mean(total for _, total in monthly_variable[index - 11:index + 1])
                if trailing_average != 0:
                    seasonal_ratios.append(month_total / trailing_average)
            if history_months >= 12 and seasonal_ratios:
                seasonal_months_observed = len(seasonal_ratios)
                raw_factor = _percentile(seasonal_ratios, 0.5)
                shrink = seasonal_months_observed / (seasonal_months_observed + 1)
                seasonal_factor = min(max(1 + (raw_factor - 1) * shrink, 0.5), 2.0)

        adjusted_weekday_rates = {p: rate * seasonal_factor for p, rate in weekday_rates.items()}
        adjusted_weekend_rates = {p: rate * seasonal_factor for p, rate in weekend_rates.items()}
        rate_remaining = {
            p: adjusted_weekday_rates[p] * weekdays_remaining
            + adjusted_weekend_rates[p] * weekend_days_remaining
            for p in (0.25, 0.5, 0.75)
        }

        if basis == "profile":
            typical_total = _percentile(historical_totals, 0.5)
            regular = months_with_any_spend / months_of_history >= 0.8
            # The floor sees only variable history and variable current spend. Recognised
            # commitments have already been removed, so rent cannot enter both paths.
            floor_remaining = (
                max(typical_total - variable_spent, 0.0)
                if regular and variable_spent == 0 and committed_remaining == 0
                else 0.0
            )
            # Adding weekday and weekend percentiles is deliberately conservative: the
            # remaining days will not all land at their respective 75th percentiles.
            variable_25 = max(rate_remaining[0.25], floor_remaining * 0.9)
            variable_50 = max(rate_remaining[0.5], floor_remaining)
            variable_75 = max(rate_remaining[0.75], floor_remaining * 1.1)
        else:
            variable_25 = variable_50 = variable_75 = variable_spent / day_of_month * len(future_dates)

        variable_low = min(variable_25, variable_50, variable_75)
        variable_medium = variable_50
        variable_high = max(variable_25, variable_50, variable_75)
        estimate_25 = spent + committed_remaining + variable_low
        estimate_50 = spent + committed_remaining + variable_medium
        estimate_75 = spent + committed_remaining + variable_high

        # Keep the median estimate intact while clamping the outer bounds around it. This
        # protects the public low <= medium <= high contract if floor and percentile
        # behavior ever cross for an unusual signed category.
        low = round(min(estimate_25, estimate_50, estimate_75), 2)
        medium = round(estimate_50, 2)
        high = round(max(estimate_25, estimate_50, estimate_75), 2)
        net_negative = spent < 0 or estimate_50 < 0

        actual_by_day = defaultdict(float)
        for tx in current[category]:
            actual_by_day[int(tx.date[-2:])] += tx.cost
        running = 0.0
        cumulative_rows = []
        for day in range(1, days_in_month + 1):
            running += actual_by_day[day]
            actual = round(running, 2) if day <= day_of_month else None
            if day < day_of_month:
                projected = {"committed": None, "low": None, "medium": None, "high": None}
            elif day == day_of_month:
                projected = {"committed": spent, "low": spent, "medium": spent, "high": spent}
            else:
                projection_dates = [today + timedelta(days=offset) for offset in range(1, day - day_of_month + 1)]
                projected_weekdays = sum(projected_day.weekday() < 5 for projected_day in projection_dates)
                projected_weekends = len(projection_dates) - projected_weekdays

                def variable_to_day(percentile: float, endpoint: float) -> float:
                    raw_endpoint = rate_remaining[percentile]
                    raw_to_day = (
                        adjusted_weekday_rates[percentile] * projected_weekdays
                        + adjusted_weekend_rates[percentile] * projected_weekends
                    )
                    if raw_endpoint:
                        return raw_to_day * endpoint / raw_endpoint
                    return endpoint * len(projection_dates) / max(len(future_dates), 1)

                committed_to_day = sum(
                    item["expected_amount"]
                    for item in committed_items
                    if item["category"] == category
                    and item["status"] != "charged"
                    and min(days_in_month, max(item["expected_day"], day_of_month + 1)) <= day
                )
                projected = {
                    "committed": round(spent + committed_to_day, 2),
                    "low": round(spent + committed_to_day + variable_to_day(0.25, variable_low), 2),
                    "medium": round(spent + committed_to_day + variable_to_day(0.5, variable_medium), 2),
                    "high": round(spent + committed_to_day + variable_to_day(0.75, variable_high), 2),
                }
            cumulative_rows.append({"day": day, "actual": actual, **projected})

        category_rows.append({
            "category": category,
            "label": labels.get(category, category_label(category)),
            "spent": spent,
            "low": low,
            "medium": medium,
            "high": high,
            "committed": round(committed_remaining, 2),
            "variable_low": round(variable_low, 2),
            "variable_medium": round(variable_medium, 2),
            "variable_high": round(variable_high, 2),
            "weekday_rate": round(adjusted_weekday_rates[0.5], 2),
            "weekend_rate": round(adjusted_weekend_rates[0.5], 2),
            "seasonal_factor": round(seasonal_factor, 4),
            "seasonal_months_observed": seasonal_months_observed,
            "basis": basis,
            "months_of_history": months_of_history,
            "net_negative": net_negative,
            "cumulative": cumulative_rows,
        })

    category_rows.sort(key=lambda row: (-row["spent"], row["category"]))
    total_spent = round(sum(row["spent"] for row in category_rows), 2)
    total_low = round(sum(row["low"] for row in category_rows), 2)
    total_medium = round(sum(row["medium"] for row in category_rows), 2)
    total_high = round(sum(row["high"] for row in category_rows), 2)
    total_cumulative = []
    for day in range(1, days_in_month + 1):
        def total_for(key: str) -> float | None:
            values = [row["cumulative"][day - 1][key] for row in category_rows]
            if not values:
                if key == "actual":
                    return 0.0 if day <= day_of_month else None
                return 0.0 if day >= day_of_month else None
            return None if all(value is None for value in values) else round(sum(value or 0 for value in values), 2)

        total_cumulative.append({
            "day": day,
            "actual": total_for("actual"),
            "committed": total_for("committed"),
            "low": total_for("low"),
            "medium": total_for("medium"),
            "high": total_for("high"),
        })

    return {
        "month": current_month,
        "day_of_month": day_of_month,
        "days_in_month": days_in_month,
        "total": {
            "spent": total_spent,
            "low": total_low,
            "medium": total_medium,
            "high": total_high,
            "variable_low": round(sum(row["variable_low"] for row in category_rows), 2),
            "variable_medium": round(sum(row["variable_medium"] for row in category_rows), 2),
            "variable_high": round(sum(row["variable_high"] for row in category_rows), 2),
            "basis": "profile" if any(row["basis"] == "profile" for row in category_rows) else "run_rate",
        },
        "committed": {
            "remaining": round(sum(committed_by_category.values()), 2),
            "charged_so_far": round(sum(committed_charged_by_category.values()), 2),
            "variable_excluded_from_history": variable_excluded_from_history,
            "items": committed_items,
        },
        "calendar": {
            "weekdays_remaining": weekdays_remaining,
            "weekend_days_remaining": weekend_days_remaining,
            "today_is_weekend": today.weekday() >= 5,
        },
        "categories": category_rows,
        "cumulative": total_cumulative,
    }


def _forecast_day_bucket(day_of_month: int) -> str:
    if day_of_month <= 10:
        return "1-10"
    if day_of_month <= 20:
        return "11-20"
    return "21-end"


def _forecast_accuracy_stats(
    transactions: list[Transaction],
    snapshots: list[dict[str, Any]],
    today: date,
) -> dict[tuple[str, str], dict[str, Any]]:
    """Score only closed months, grouped by category and prediction-time bucket."""
    current_month = month_key(today)
    actual_by_month_category: dict[tuple[str, str], float] = defaultdict(float)
    actual_by_month: dict[str, float] = defaultdict(float)
    for tx in transactions:
        if tx.month >= current_month or not tx_is_expense(tx):
            continue
        actual_by_month_category[(tx.month, tx.category)] += tx.cost
        actual_by_month[tx.month] += tx.cost

    scored: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for snapshot in snapshots:
        if (
            snapshot.get("model_version") != FORECAST_MODEL_VERSION
            or str(snapshot.get("month", "")) >= current_month
        ):
            continue
        month = str(snapshot["month"])
        category = str(snapshot.get("category", ""))
        actual = (
            actual_by_month.get(month)
            if category == ""
            else actual_by_month_category.get((month, category))
        )
        # Match the SQL inner join: a snapshot without a final category total is not a
        # scored observation. The month total is present whenever that month has spend.
        if actual is None:
            continue
        medium = float(snapshot["medium"])
        pct_error = (actual - medium) / abs(actual) if actual else None
        bucket = _forecast_day_bucket(int(snapshot["day_of_month"]))
        scored[(category, bucket)].append({
            "month": month,
            "day_of_month": int(snapshot["day_of_month"]),
            "predicted_medium": medium,
            "actual": actual,
            "pct_error": pct_error,
            "in_band": float(snapshot["low"]) <= actual <= float(snapshot["high"]),
        })

    stats: dict[tuple[str, str], dict[str, Any]] = {}
    for key, rows in scored.items():
        pct_errors = [row["pct_error"] for row in rows if row["pct_error"] is not None]
        stats[key] = {
            "sample_count": len(rows),
            "months": len({row["month"] for row in rows}),
            "median_pct_error": _percentile(pct_errors, 0.5) if pct_errors else None,
            "in_band_rate": mean(1.0 if row["in_band"] else 0.0 for row in rows),
            "rows": rows,
        }
    return stats


def build_forecast(
    transactions: list[Transaction],
    forecast_snapshots: list[dict[str, Any]] | None = None,
) -> dict[str, Any]:
    """Add closed-month scoring and bounded calibration to the day-type forecast.

    Snapshot persistence belongs to the cloud's volatile capture function. Accepting
    snapshot rows here keeps the local Python mirror deterministic and makes the scoring
    and correction maths independently testable; ordinary local builds have no cloud
    snapshots and therefore leave every forecast number unchanged.
    """
    forecast = _build_forecast_daytype_base(transactions)
    snapshots = forecast_snapshots or []
    today = date.today()
    bucket = _forecast_day_bucket(forecast["day_of_month"])
    stats = _forecast_accuracy_stats(transactions, snapshots, today)
    corrections_applied = False

    for row in forecast["categories"]:
        category_stats = stats.get((row["category"], bucket))
        months = int(category_stats["months"]) if category_stats else 0
        sample_count = int(category_stats["sample_count"]) if category_stats else 0
        median_pct_error = category_stats["median_pct_error"] if category_stats else None
        in_band_rate = category_stats["in_band_rate"] if category_stats else None

        if months < 2:
            bias_factor = 1.0
            band_scale = 1.0
        else:
            shrink = months / (months + 3)
            bias_factor = min(max(1 + (median_pct_error or 0.0) * shrink, 0.7), 1.4)
            band_scale = min(max(math.sqrt(0.5 / max(in_band_rate or 0.0, 0.1)), 0.6), 1.8)
            corrections_applied = True

        row["accuracy"] = {
            "median_pct_error": median_pct_error,
            "in_band_rate": in_band_rate,
            "bias_factor": bias_factor,
            "band_scale": band_scale,
            "months": months,
            "sample_count": sample_count,
            "day_bucket": bucket,
        }

        # Do not even reconstruct the existing values below the learning threshold.
        # This guarantees the Addendum 4 numerical payload remains identical today.
        if months < 2:
            continue

        variable_low = float(row["variable_low"])
        variable_medium = float(row["variable_medium"])
        variable_high = float(row["variable_high"])
        corrected_medium = variable_medium * bias_factor
        corrected_low = min(
            corrected_medium - (variable_medium - variable_low) * band_scale,
            corrected_medium,
        )
        corrected_high = max(
            corrected_medium + (variable_high - variable_medium) * band_scale,
            corrected_medium,
        )
        row["variable_low"] = round(corrected_low, 2)
        row["variable_medium"] = round(corrected_medium, 2)
        row["variable_high"] = round(corrected_high, 2)
        row["low"] = round(row["spent"] + row["committed"] + corrected_low, 2)
        row["medium"] = round(row["spent"] + row["committed"] + corrected_medium, 2)
        row["high"] = round(row["spent"] + row["committed"] + corrected_high, 2)
        row["net_negative"] = row["spent"] < 0 or row["medium"] < 0

        for point in row["cumulative"]:
            if point["day"] <= forecast["day_of_month"]:
                continue
            committed_line = float(point["committed"])
            point_low = float(point["low"]) - committed_line
            point_medium = float(point["medium"]) - committed_line
            point_high = float(point["high"]) - committed_line
            corrected_point_medium = point_medium * bias_factor
            corrected_point_low = min(
                corrected_point_medium - (point_medium - point_low) * band_scale,
                corrected_point_medium,
            )
            corrected_point_high = max(
                corrected_point_medium + (point_high - point_medium) * band_scale,
                corrected_point_medium,
            )
            point["low"] = round(committed_line + corrected_point_low, 2)
            point["medium"] = round(committed_line + corrected_point_medium, 2)
            point["high"] = round(committed_line + corrected_point_high, 2)

    if corrections_applied:
        categories = forecast["categories"]
        forecast["total"].update({
            "spent": round(sum(row["spent"] for row in categories), 2),
            "low": round(sum(row["low"] for row in categories), 2),
            "medium": round(sum(row["medium"] for row in categories), 2),
            "high": round(sum(row["high"] for row in categories), 2),
            "variable_low": round(sum(row["variable_low"] for row in categories), 2),
            "variable_medium": round(sum(row["variable_medium"] for row in categories), 2),
            "variable_high": round(sum(row["variable_high"] for row in categories), 2),
        })
        for day_index, total_point in enumerate(forecast["cumulative"]):
            for key in ("actual", "committed", "low", "medium", "high"):
                values = [row["cumulative"][day_index][key] for row in categories]
                total_point[key] = (
                    None
                    if all(value is None for value in values)
                    else round(sum(value or 0 for value in values), 2)
                )

    total_stats = stats.get(("", bucket))
    total_months = int(total_stats["months"]) if total_stats else 0
    # The SQL chooses the latest captured day in the latest completed forecast month.
    # Build this directly from the total rows so other category stats cannot enter.
    last_rows = [
        row
        for (category, _), value in stats.items()
        if category == ""
        for row in value["rows"]
    ]
    last_row = max(last_rows, key=lambda row: (row["month"], row["day_of_month"]), default=None)
    last_month = None if last_row is None else {
        "month": last_row["month"],
        "predicted_medium": round(last_row["predicted_medium"], 2),
        "actual": round(last_row["actual"], 2),
        "pct_error": last_row["pct_error"],
        "in_band": last_row["in_band"],
    }
    forecast["accuracy"] = {
        "model_version": FORECAST_MODEL_VERSION,
        "months_scored": total_months,
        "total": {
            "median_pct_error": total_stats["median_pct_error"] if total_stats else None,
            "in_band_rate": total_stats["in_band_rate"] if total_stats else None,
        },
        "corrections_applied": corrections_applied,
        "by_category": [
            {
                "category": row["category"],
                "median_pct_error": row["accuracy"]["median_pct_error"],
                "in_band_rate": row["accuracy"]["in_band_rate"],
                "bias_factor": row["accuracy"]["bias_factor"],
                "band_scale": row["accuracy"]["band_scale"],
                "months": row["accuracy"]["months"],
            }
            for row in sorted(forecast["categories"], key=lambda value: value["category"])
        ],
        "last_month": last_month,
    }
    return forecast


def parse_workbook_monthly_summary(wb: openpyxl.Workbook) -> dict[str, Any]:
    ws = wb["Monthly Summary"]
    header = [ws.cell(4, col).value for col in range(3, 18)]
    months = []
    for value in header:
        d = parse_date(value)
        if d:
            months.append(month_key(d))

    row_specs = [
        (5, "normal", "money"),
        (6, "normal", "money"),
        (7, "normal", "money"),
        (8, "normal", "money"),
        (9, "normal", "money"),
        (10, "normal", "money"),
        (11, "normal", "money"),
        (12, "normal", "money"),
        (13, "normal", "money"),
        (14, "normal", "money"),
        (15, "normal", "money"),
        (16, "normal", "money"),
        (17, "normal", "money"),
        (18, "normal", "money"),
        (19, "normal", "money"),
        (20, "normal", "money"),
        (21, "normal", "money"),
        (22, "normal", "money"),
        (23, "normal", "money"),
        (24, "normal", "money"),
        (26, "subtotal-row", "money"),
        (None, "spacer-row", "money"),
        (28, "normal", "money"),
        (None, "spacer-row", "money"),
        (30, "normal", "money"),
        (32, "subtotal-row", "money"),
        (33, "margin-row", "percent"),
        (None, "spacer-row", "money"),
        (35, "normal", "integer"),
        (36, "normal", "money"),
        (37, "normal", "money"),
        (38, "normal", "money"),
        (39, "subtotal-row", "money"),
        (None, "spacer-row", "money"),
        (41, "subtotal-row", "money"),
    ]
    rows = []
    for row_num, style, kind in row_specs:
        if row_num is None:
            rows.append({"label": "", "style": style, "kind": kind, "values": {}})
            continue
        label = clean_text(ws.cell(row_num, 2).value)
        values = {}
        for index, month in enumerate(months, start=3):
            values[month] = round(money(ws.cell(row_num, index).value), 6)
        rows.append({"label": label, "style": style, "kind": kind, "values": values})
    return {"months": months, "rows": rows}


def extend_monthly_summary_with_actuals(
    workbook_monthly: dict[str, Any], transactions: list[Transaction]
) -> dict[str, Any]:
    """Fill any transaction months not present in the budget workbook (e.g. months
    after the workbook's last column) with actuals computed from transactions, so the
    Monthly summary is first-class instead of runtime-derived.

    Spend definitions match the Overview (build_summary): a month's Total is the sum of
    every non-internal expense category cost except rent. ``mom`` is a contra-expense
    (reimbursements from Mom against spend booked elsewhere), so it stays in the spend
    block with a negative value and pulls Total down; ``event`` is a real expense
    category with no workbook row, so one is added here to keep the category rows
    reconciling with Total.
    """
    wb_months = set(workbook_monthly["months"])
    missing = sorted({tx.month for tx in transactions} - wb_months)
    if not missing:
        return workbook_monthly

    rows = workbook_monthly["rows"]
    if not any(row.get("label") == "event" for row in rows):
        total_idx = next((i for i, row in enumerate(rows) if row.get("label") == "Total"), len(rows))
        rows.insert(total_idx, {"label": "event", "style": "normal", "kind": "money", "values": {}})

    cost_by: dict[tuple[str, str], float] = defaultdict(float)
    for tx in transactions:
        cost_by[(tx.month, tx.category)] += tx.cost

    def cat_cost(month: str, cat: str) -> float:
        return round(cost_by.get((month, cat), 0.0), 2)

    for month in missing:
        total = round(
            sum(
                cost
                for (m, cat), cost in cost_by.items()
                if m == month and cat not in INTERNAL_CATEGORIES and cat != "rent"
            ),
            2,
        )
        rent_val = round(-cat_cost(month, "rent"), 2)
        salary_val = round(-cat_cost(month, "salary"), 2)
        k401 = 779.17 if month < "2025-07" else 830.50
        ira = 141.67 if month < "2025-07" else 151.00
        total_inv = round(k401 + ira, 2)
        net_income = round(salary_val - total + rent_val, 2)
        annual_salary = 85000.0 if month < "2025-07" else 90600.0
        special = {
            "Total": total,
            "Rent": rent_val,
            "Salary": salary_val,
            "Net Income": net_income,
            "Margin": round(net_income / salary_val, 6) if salary_val else 0.0,
            "401k": k401,
            "IRA": ira,
            "Additional Savings": 0.0,
            "Total Inv Savings": total_inv,
            "Total Savings": round(net_income + total_inv, 2),
        }
        for row in rows:
            label = row.get("label", "")
            if row.get("style") == "spacer-row":
                continue
            row.setdefault("values", {})
            if label in special:
                row["values"][month] = special[label]
            elif not label and row.get("kind") == "integer":
                row["values"][month] = annual_salary  # base annual salary row
            else:
                row["values"][month] = cat_cost(month, label.lower())

    workbook_monthly["months"] = sorted(wb_months | set(missing))
    return workbook_monthly


def reconcile_salary_rows(
    workbook_monthly: dict[str, Any], transactions: list[Transaction]
) -> dict[str, Any]:
    """Force the Monthly 'Salary' row to equal actual salary transactions so it matches
    the Overview 'Income' KPI. Corrects months where the budget workbook folded other
    inflows (tax refunds, deposited checks) into salary, and recomputes the dependent
    rows (Net Income, Margin, Total Savings) so the summary stays internally consistent.
    """
    salary_by_month: dict[str, float] = defaultdict(float)
    for tx in transactions:
        if tx.category == "salary":
            salary_by_month[tx.month] += -tx.cost

    rows = {row.get("label"): row for row in workbook_monthly["rows"]}
    salary_row = rows.get("Salary")
    if not salary_row:
        return workbook_monthly
    total_row, rent_row = rows.get("Total"), rows.get("Rent")
    net_row, margin_row = rows.get("Net Income"), rows.get("Margin")
    inv_row, save_row = rows.get("Total Inv Savings"), rows.get("Total Savings")

    def val(row: dict[str, Any] | None, month: str) -> float:
        return float(row["values"].get(month, 0) or 0) if row else 0.0

    for month in workbook_monthly["months"]:
        new_sal = round(salary_by_month.get(month, 0.0), 2)
        if abs(new_sal - round(val(salary_row, month), 2)) <= 0.02:
            continue
        salary_row["values"][month] = new_sal
        net = round(new_sal - val(total_row, month) + val(rent_row, month), 2)
        if net_row:
            net_row["values"][month] = net
        if margin_row:
            margin_row["values"][month] = round(net / new_sal, 6) if new_sal else 0.0
        if save_row:
            save_row["values"][month] = round(net + val(inv_row, month), 2)
    return workbook_monthly


def connect_database() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def migrate_database(conn: sqlite3.Connection) -> None:
    """Add columns to an existing transactions table.

    CREATE TABLE IF NOT EXISTS is a no-op once the table exists, so a new column never
    lands on a database built by an earlier version. Dropping and recreating is not an
    option: category_overrides lives in the same file and must survive.
    """
    existing = {row[1] for row in conn.execute("PRAGMA table_info(transactions)")}
    if existing and "occurrence" not in existing:
        conn.execute("ALTER TABLE transactions ADD COLUMN occurrence INTEGER NOT NULL DEFAULT 1")


def ensure_database(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS metadata (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS transactions (
            id TEXT PRIMARY KEY,
            account TEXT NOT NULL,
            date TEXT NOT NULL,
            month TEXT NOT NULL,
            description TEXT NOT NULL,
            category TEXT NOT NULL,
            category_label TEXT NOT NULL,
            amount REAL NOT NULL,
            cost REAL NOT NULL,
            type TEXT NOT NULL DEFAULT '',
            status TEXT NOT NULL DEFAULT '',
            source TEXT NOT NULL,
            native_category TEXT NOT NULL DEFAULT '',
            counterparty TEXT NOT NULL DEFAULT '',
            needs_review INTEGER NOT NULL DEFAULT 0,
            occurrence INTEGER NOT NULL DEFAULT 1
        );

        CREATE TABLE IF NOT EXISTS category_overrides (
            tx_id TEXT PRIMARY KEY,
            category TEXT NOT NULL,
            note TEXT NOT NULL DEFAULT '',
            updated_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS monthly_summary_rows (
            row_order INTEGER NOT NULL,
            label TEXT NOT NULL DEFAULT '',
            style TEXT NOT NULL DEFAULT '',
            kind TEXT NOT NULL DEFAULT 'money',
            month TEXT NOT NULL,
            value REAL,
            PRIMARY KEY (row_order, month)
        );

        CREATE INDEX IF NOT EXISTS idx_transactions_month ON transactions(month);
        CREATE INDEX IF NOT EXISTS idx_transactions_account_month ON transactions(account, month);
        CREATE INDEX IF NOT EXISTS idx_transactions_category_month ON transactions(category, month);
        CREATE INDEX IF NOT EXISTS idx_transactions_needs_review ON transactions(needs_review);

        CREATE VIEW IF NOT EXISTS v_account_month_category AS
            SELECT
                account,
                month,
                category,
                category_label,
                COUNT(*) AS transaction_count,
                ROUND(SUM(amount), 2) AS amount_total,
                ROUND(SUM(cost), 2) AS cost_total
            FROM transactions
            GROUP BY account, month, category, category_label;

        CREATE VIEW IF NOT EXISTS v_month_category AS
            SELECT
                month,
                category,
                category_label,
                COUNT(*) AS transaction_count,
                ROUND(SUM(amount), 2) AS amount_total,
                ROUND(SUM(cost), 2) AS cost_total
            FROM transactions
            GROUP BY month, category, category_label;
        """
    )
    migrate_database(conn)


def needs_review(tx: Transaction) -> bool:
    if tx.category in {"uncategorized", "merchandise"}:
        return True
    return bool(tx.needs_review)


def load_db_category_overrides(conn: sqlite3.Connection) -> dict[str, str]:
    ensure_database(conn)
    return {row["tx_id"]: row["category"] for row in conn.execute("SELECT tx_id, category FROM category_overrides")}


def apply_db_category_overrides(transactions: list[Transaction], overrides: dict[str, str]) -> None:
    for tx in transactions:
        category = overrides.get(tx.id)
        if not category:
            continue
        tx.category = canonical_category(category)
        tx.category_label = category_label(tx.category)


def write_database(conn: sqlite3.Connection, transactions: list[Transaction], monthly_summary: dict[str, Any]) -> None:
    ensure_database(conn)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with conn:
        conn.execute("DELETE FROM transactions")
        conn.execute("DELETE FROM monthly_summary_rows")
        conn.execute("DELETE FROM metadata")
        conn.executemany(
            """
            INSERT INTO transactions (
                id, account, date, month, description, category, category_label, amount, cost,
                type, status, source, native_category, counterparty, needs_review, occurrence
            )
            VALUES (
                :id, :account, :date, :month, :description, :category, :category_label, :amount, :cost,
                :type, :status, :source, :native_category, :counterparty, :needs_review, :occurrence
            )
            """,
            [
                {
                    **asdict(tx),
                    "needs_review": 1 if needs_review(tx) else 0,
                }
                for tx in transactions
            ],
        )
        monthly_rows = []
        for row_order, row in enumerate(monthly_summary["rows"]):
            values = row.get("values") or {}
            if not values:
                monthly_rows.append(
                    {
                        "row_order": row_order,
                        "label": row.get("label", ""),
                        "style": row.get("style", ""),
                        "kind": row.get("kind", "money"),
                        "month": "",
                        "value": None,
                    }
                )
                continue
            for month, value in values.items():
                monthly_rows.append(
                    {
                        "row_order": row_order,
                        "label": row.get("label", ""),
                        "style": row.get("style", ""),
                        "kind": row.get("kind", "money"),
                        "month": month,
                        "value": value,
                    }
                )
        conn.executemany(
            """
            INSERT INTO monthly_summary_rows (row_order, label, style, kind, month, value)
            VALUES (:row_order, :label, :style, :kind, :month, :value)
            """,
            monthly_rows,
        )
        conn.executemany(
            "INSERT INTO metadata (key, value) VALUES (?, ?)",
            [
                ("generated_at", generated_at),
                ("source_workbook", str(WORKBOOK)),
                ("transaction_count", str(len(transactions))),
                ("monthly_summary_months", json.dumps(monthly_summary["months"])),
            ],
        )


def load_transactions_from_database(conn: sqlite3.Connection) -> list[Transaction]:
    rows = conn.execute(
        """
        SELECT id, account, date, month, description, category, category_label, amount, cost,
               type, status, source, native_category, counterparty, needs_review, occurrence
        FROM transactions
        ORDER BY date DESC, account, description
        """
    ).fetchall()
    return [Transaction(**dict(row)) for row in rows]


def load_monthly_summary_from_database(conn: sqlite3.Connection) -> dict[str, Any]:
    month_rows = conn.execute(
        "SELECT value FROM metadata WHERE key = 'monthly_summary_months'"
    ).fetchone()
    months = json.loads(month_rows["value"]) if month_rows else []
    rows_by_order: dict[int, dict[str, Any]] = {}
    for row in conn.execute(
        """
        SELECT row_order, label, style, kind, month, value
        FROM monthly_summary_rows
        ORDER BY row_order, month
        """
    ):
        item = rows_by_order.setdefault(
            row["row_order"],
            {"label": row["label"], "style": row["style"], "kind": row["kind"], "values": {}},
        )
        if row["month"]:
            item["values"][row["month"]] = row["value"]
    return {"months": months, "rows": [rows_by_order[key] for key in sorted(rows_by_order)]}


def build_insights(transactions: list[Transaction], summary: dict[str, Any]) -> list[dict[str, str]]:
    insights: list[dict[str, str]] = []
    months = [row["month"] for row in summary["months"]]
    complete_months = [m for m in months if m < months[-1]] if months else []
    recent_months = complete_months[-3:]
    prior_months = complete_months[:-3]

    def category_month_spend(category: str, month: str) -> float:
        return sum(tx.cost for tx in transactions if tx.category == category and tx.month == month and tx_is_expense(tx))

    variable_categories = [row for row in summary["categories"] if row["category"] not in {"rent"} and row["expense"] > 0]
    for row in variable_categories[:5]:
        insights.append(
            {
                "title": f"Watch {row['label']}",
                "detail": f"{row['label']} is ${row['expense']:,.0f} across the loaded period across {row['count']} transactions.",
                "action": "Review the largest merchants in this category before setting next month's budget.",
                "tone": "neutral",
            }
        )

    if recent_months and prior_months:
        rising = []
        for row in variable_categories:
            recent = mean([category_month_spend(row["category"], m) for m in recent_months])
            prior = mean([category_month_spend(row["category"], m) for m in prior_months])
            if prior > 50 and recent > prior * 1.25:
                rising.append((recent - prior, row["label"], recent, prior))
        for diff, label, recent, prior in sorted(rising, reverse=True)[:3]:
            insights.append(
                {
                    "title": f"{label} is trending up",
                    "detail": f"The recent monthly average is ${recent:,.0f}, up ${diff:,.0f} from the earlier average of ${prior:,.0f}.",
                    "action": "Set a category cap or inspect repeat merchants for this category.",
                    "tone": "warning",
                }
            )

    recurring = []
    groups: dict[tuple[str, int], list[Transaction]] = defaultdict(list)
    for tx in transactions:
        if not tx_is_expense(tx) or tx.cost <= 0:
            continue
        merchant = normalize_merchant(tx.description)
        if not merchant:
            continue
        groups[(merchant, round(tx.cost))].append(tx)
    for (merchant, rounded_cost), rows in groups.items():
        months_seen = {tx.month for tx in rows}
        if len(rows) >= 3 and len(months_seen) >= 3 and rounded_cost >= 3:
            recurring.append((len(months_seen), rounded_cost, merchant.title(), rows))
    for _, rounded_cost, merchant, rows in sorted(recurring, reverse=True)[:5]:
        insights.append(
            {
                "title": f"Recurring charge: {merchant}",
                "detail": f"Appears in {len({tx.month for tx in rows})} months at roughly ${rounded_cost:,.0f}.",
                "action": "Keep it only if it is still intentional; otherwise cancel or downgrade it.",
                "tone": "opportunity",
            }
        )

    betting = sum(tx.cost for tx in transactions if tx.category == "bet" and tx_is_expense(tx))
    if betting > 0:
        insights.append(
            {
                "title": "Betting spend is visible",
                "detail": f"Loaded betting-related spend nets to ${betting:,.0f}.",
                "action": "Use a hard monthly limit because this category can move quickly and does not compound value.",
                "tone": "warning",
            }
        )

    service = sum(tx.cost for tx in transactions if tx.category in {"service", "spotify"} and tx_is_expense(tx))
    if service > 0:
        insights.append(
            {
                "title": "Audit services and subscriptions",
                "detail": f"Services and subscriptions total ${service:,.0f} in the loaded data.",
                "action": "Cancel unused software, media, and marketplace subscriptions; this is the easiest recurring reduction area.",
                "tone": "opportunity",
            }
        )

    return insights[:12]


def build_cashflow(transactions: list[Transaction]) -> dict[str, Any]:
    """Model the checking balance as a stock driven by monthly flows.

    Spend and cash are different questions and they disagree every month. Spend is booked
    when a card is charged; cash moves when that card bill is paid, often a month later.
    The bridge below reconciles the two exactly, so a month where the P&L looks fine but
    the balance dropped has a visible, itemised reason rather than looking like an error.
    """
    months = sorted({tx.month for tx in transactions})
    anchor_month = OPENING_BALANCE_ANCHOR["as_of"].strftime("%Y-%m")

    cash_rows = [tx for tx in transactions if tx.account == CASH_ACCOUNT]
    by_month: dict[str, list[Transaction]] = defaultdict(list)
    for tx in cash_rows:
        by_month[tx.month].append(tx)

    def flow(month: str, predicate) -> float:
        return round(sum(tx.amount for tx in by_month.get(month, []) if predicate(tx)), 2)

    # Walk the anchor outward. Balances before the anchor rely on workbook-era checking
    # rows, which were entered for budgeting rather than as a full ledger, so they are
    # marked derived-only and the UI says so rather than implying bank-grade accuracy.
    delta = {m: round(sum(tx.amount for tx in by_month.get(m, [])), 2) for m in months}
    closing: dict[str, float] = {}
    if anchor_month in months:
        closing[anchor_month] = round(OPENING_BALANCE_ANCHOR["balance"], 2)
        idx = months.index(anchor_month)
        for m in months[idx + 1:]:
            closing[m] = round(closing[months[months.index(m) - 1]] + delta[m], 2)
        for m in reversed(months[:idx]):
            nxt = months[months.index(m) + 1]
            closing[m] = round(closing[nxt] - delta[nxt], 2)

    out: list[dict[str, Any]] = []
    for month in months:
        rows = [tx for tx in transactions if tx.month == month]
        salary_in = flow(month, lambda tx: tx.category in INCOME_CATEGORIES)
        card_pmts = flow(month, lambda tx: tx.category == "payment")
        rent_out = flow(month, lambda tx: tx.category == "rent")
        invest_out = flow(month, lambda tx: tx.category in SAVINGS_CATEGORIES)
        transfers_in = flow(month, lambda tx: tx.category in TRANSFER_CATEGORIES and tx.category != "payment")
        other_cash = round(
            delta[month] - salary_in - card_pmts - rent_out - invest_out - transfers_in, 2
        )

        # Accrual side, measured across every account.
        income = round(sum(income_amount(tx) for tx in rows), 2)
        spend = round(sum(spend_cost(tx) for tx in rows), 2)
        net_income = round(income - spend, 2)
        # Spend that was charged somewhere other than checking has not touched cash yet.
        deferred = round(
            sum(spend_cost(tx) for tx in rows if tx.account != CASH_ACCOUNT), 2
        )
        bridge_residual = round(
            delta[month] - (net_income + deferred + card_pmts + invest_out + transfers_in), 2
        )

        # Card float: charged but not yet paid. Only the change is trustworthy -- the
        # opening float is unknown -- so it is reported as a running total since the start.
        charged = round(
            sum(spend_cost(tx) for tx in rows if tx.account in CARD_ACCOUNTS), 2
        )
        paid = round(-card_pmts, 2)

        out.append({
            "month": month,
            "opening": round(closing[month] - delta[month], 2) if month in closing else None,
            "closing": closing.get(month),
            "delta": delta[month],
            "salary_in": salary_in,
            "transfers_in": transfers_in,
            "card_payments": card_pmts,
            "rent": rent_out,
            "investments": invest_out,
            "other_cash": other_cash,
            "net_income": net_income,
            "income": income,
            "spend": spend,
            "deferred_spend": deferred,
            "bridge_residual": bridge_residual,
            "charged_to_cards": charged,
            "paid_to_cards": paid,
            "float_change": round(charged - paid, 2),
        })

    running = 0.0
    for row in out:
        running = round(running + row["float_change"], 2)
        row["float_cumulative"] = running

    return {
        "anchor": {
            "account": OPENING_BALANCE_ANCHOR["account"],
            "as_of": OPENING_BALANCE_ANCHOR["as_of"].isoformat(),
            "balance": round(OPENING_BALANCE_ANCHOR["balance"], 2),
            "month": anchor_month,
        },
        "months": out,
    }


def build_payload() -> dict[str, Any]:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True, read_only=True)
    mapping = build_mapping(wb)
    transactions = parse_workbook_transactions(wb) + parse_recent_csvs(mapping)
    plaid_transactions = parse_plaid_transactions(mapping)
    if plaid_transactions:
        transactions, superseded = apply_plaid_takeover(transactions + plaid_transactions)
        print(f"Plaid: {len(plaid_transactions)} rows from {PLAID_TAKEOVER_DATE}, "
              f"superseding {superseded} CSV rows in {', '.join(sorted(plaid_covered_accounts()))}")
    workbook_monthly = parse_workbook_monthly_summary(wb)
    seen = set()
    deduped = []
    for tx in transactions:
        if tx.source.startswith("Workbook:"):
            fingerprint = (tx.source, tx.account, tx.date, tx.description.lower(), round(tx.amount, 2))
        else:
            # counterparty distinguishes four roommates each sending the same amount with
            # the same note on the same day; occurrence distinguishes genuine repeat
            # charges (two Metro taps) from one charge appearing in two overlapping
            # exports. Both were previously collapsed into a single row.
            fingerprint = (
                tx.account,
                tx.date,
                tx.description.lower(),
                round(tx.amount, 2),
                tx.source.split(":")[0],
                tx.counterparty.lower(),
                tx.occurrence,
            )
        if fingerprint in seen:
            continue
        seen.add(fingerprint)
        deduped.append(tx)
    deduped.sort(key=lambda tx: (tx.date, tx.account, tx.description), reverse=True)

    with connect_database() as conn:
        ensure_database(conn)
        apply_db_category_overrides(deduped, load_db_category_overrides(conn))
        extend_monthly_summary_with_actuals(workbook_monthly, deduped)
        reconcile_salary_rows(workbook_monthly, deduped)
        write_database(conn, deduped, workbook_monthly)
        db_transactions = load_transactions_from_database(conn)
        db_workbook_monthly = load_monthly_summary_from_database(conn)

    summary = build_summary(db_transactions)
    return {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "database": str(DB_PATH),
        "transactions": [asdict(tx) for tx in db_transactions],
        "summary": summary,
        "workbook_monthly": db_workbook_monthly,
        "insights": build_insights(db_transactions, summary),
        "cashflow": build_cashflow(db_transactions),
        "forecast": build_forecast(db_transactions),
        "accounts": ["Overview", "Forecast", "Monthly", "Cash Flow", "Analytics", "Review"] + sorted({tx.account for tx in db_transactions}),
    }


def render_html(payload: dict[str, Any]) -> str:
    data = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    data = data.replace("<", "\\u003c").replace(">", "\\u003e").replace("&", "\\u0026")
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Financial Dashboard</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=Playfair+Display:wght@600;700&display=swap">
  <style>
    :root {{
      --bg: #eae9e0;
      --panel: #f4f4ec;
      --panel-strong: #f9f9f2;
      --ink: #111811;
      --muted: #536050;
      --line: rgba(28,42,30,0.13);
      --accent: #2e6b3f;
      --green-bright: #56bb6e;
      --green-soft: #dde9cd;
      --gold: #9e7132;
      --accent-2: #9e7132;
      --accent-3: #2e6b3f;
      --danger: #b13d3d;
      --good: #2e6b3f;
      --soft: #dde9cd;
      --radius: 4px;
      --font-display: 'Playfair Display', Georgia, serif;
      --font-body: Inter, ui-sans-serif, system-ui, sans-serif;
      --shadow: 0 1px 2px rgba(16,26,18,.05), 0 2px 8px rgba(16,26,18,.04);
      --shadow-md: 0 4px 20px rgba(16,26,18,.09), 0 1px 3px rgba(16,26,18,.05);
      --shadow-card: 0 4px 20px rgba(16,26,18,.09), 0 1px 3px rgba(16,26,18,.05);
      --cat-food: #c2603a;
      --cat-grocery: #4e8b5a;
      --cat-rent: #2c4a73;
      --cat-travel: #345995;
      --cat-bar: #b8893a;
      --cat-utilities: #5b8aa6;
      --cat-service: #7b5ea7;
      --cat-living: #8a7a5c;
      --cat-uber: #2f6f6a;
      --cat-clothing: #b0568a;
      --cat-gift: #c75d6f;
      --cat-fee: #9a6b4f;
      --cat-donations: #5a9e8f;
      --cat-golf: #6f9b3a;
      --cat-alc: #a85d8c;
      --cat-haircut: #7e8b3f;
      --cat-bet: #b13d3d;
      --cat-cash: #6b7a72;
      --cat-metro: #4a7aa0;
      --cat-spotify: #3d9956;
      --cat-event: #c06a2c;
      --cat-salary: #28724f;
      --cat-check: #3f7a5c;
      --cat-mom: #5e8c7d;
      --cat-stock: #436f9e;
      --cat-payment: #9aa3a0;
      --cat-venmo: #8f9aa0;
      --cat-uncategorized: #9aa3a0;
      --cat-default: #8c948f;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      font-family: var(--font-body);
      color: var(--ink);
      background: var(--bg);
    }}
    header {{
      display: grid;
      grid-template-columns: 1fr auto;
      gap: 24px;
      align-items: end;
      padding: 28px clamp(18px, 4vw, 48px) 18px;
      background: var(--panel);
      border-bottom: 1px solid var(--line);
    }}
    h1, h2, h3 {{ font-family: var(--font-display); }}
    h1 {{ margin: 0; font-size: clamp(26px, 4vw, 44px); line-height: 1.05; letter-spacing: 0; }}
    h2 {{ margin: 0 0 14px; font-size: 20px; letter-spacing: 0; }}
    h3 {{ margin: 0 0 6px; font-size: 15px; letter-spacing: 0; }}
    .subtle {{ color: var(--muted); font-size: 14px; margin-top: 8px; }}
    .toolbar {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
      align-items: center;
      justify-content: flex-end;
    }}
    input, select, button {{
      height: 38px;
      border: 1px solid var(--line);
      background: var(--panel-strong);
      color: var(--ink);
      border-radius: var(--radius);
      padding: 0 12px;
      font: inherit;
      font-size: 14px;
    }}
    button {{
      cursor: pointer;
      background: var(--soft);
      border-color: var(--line);
      color: var(--accent);
      font-weight: 700;
    }}
    nav {{
      display: flex;
      gap: 8px;
      overflow-x: auto;
      padding: 12px clamp(18px, 4vw, 48px);
      background: var(--panel);
      border-bottom: 1px solid var(--line);
    }}
    nav button {{
      flex: 0 0 auto;
      background: var(--panel);
      color: var(--muted);
      min-width: 112px;
    }}
    nav button.active {{
      background: var(--accent);
      border-color: var(--accent);
      color: #fff;
    }}
    main {{ padding: 24px clamp(18px, 4vw, 48px) 44px; }}
    .grid {{ display: grid; gap: 16px; }}
    .kpis {{ grid-template-columns: repeat(6, minmax(150px, 1fr)); }}
    .two {{ grid-template-columns: minmax(0, 1.4fr) minmax(320px, .8fr); }}
    .three {{ grid-template-columns: repeat(3, minmax(0, 1fr)); }}
    .card {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: var(--radius);
      box-shadow: var(--shadow-card);
      padding: 16px;
    }}
    .kpis .card {{ border-left: 3px solid var(--soft); }}
    .swatch {{ display: inline-block; width: 10px; height: 10px; border-radius: 3px; margin-right: 8px; vertical-align: baseline; }}
    .metric-label {{ color: var(--muted); font-size: 12px; text-transform: uppercase; font-weight: 800; letter-spacing: .06em; }}
    .metric-value {{ font-size: 26px; font-weight: 850; margin-top: 6px; line-height: 1.1; }}
    .metric-note {{ color: var(--muted); font-size: 12px; margin-top: 6px; }}
    .chart {{ width: 100%; min-height: 280px; }}
    .chart circle.pt {{ transition: r .1s; }}
    .chart .ptg:hover circle.pt {{ r: 7; }}
    #chartTip {{ position: fixed; pointer-events: none; opacity: 0; transition: opacity .12s; background: var(--panel-strong); border: 1px solid var(--line); border-radius: var(--radius); box-shadow: var(--shadow-md); padding: 8px 10px; font: 12px/1.4 var(--font-body); color: var(--ink); max-width: 240px; z-index: 50; }}
    #chartTip.show {{ opacity: 1; }}
    /* minmax(0,1fr) stops a wide child (the expanded rollup table) from stretching the
       implicit column and overflowing the card. */
    .bars {{ display: grid; gap: 10px; grid-template-columns: minmax(0, 1fr); }}
    .bar-row {{ display: grid; grid-template-columns: 130px 1fr 92px; gap: 10px; align-items: center; font-size: 13px; padding: 2px 4px; transition: background .1s; }}
    .bar-row:hover {{ background: var(--green-soft); border-radius: var(--radius); }}
    .bar-row:hover > div:first-child {{ font-weight: 600; }}
    .bar-track {{ height: 11px; background: #edf0eb; border-radius: 999px; overflow: hidden; }}
    /* Credits (negative net, e.g. mom reimbursements) read as hatched, not solid spend. */
    .bar-fill.credit {{ opacity: .5; background-image: repeating-linear-gradient(45deg, rgba(255,255,255,.65) 0 3px, transparent 3px 6px); }}
    .rollup-row {{ cursor: pointer; }}
    .rollup-row:focus-visible {{ outline: 2px solid var(--green); outline-offset: 2px; border-radius: var(--radius); }}
    .rollup-row[aria-expanded="true"] > div:first-child {{ font-weight: 600; }}
    .rollup-caret {{ display: inline-block; width: 12px; color: var(--muted); }}
    .rollup-detail {{ margin: 2px 0 10px; padding: 8px 10px; background: var(--green-soft); border-radius: var(--radius); }}
    .rollup-detail .table-wrap {{ max-height: 320px; overflow-y: auto; overflow-x: auto; }}
    /* Fixed layout so Cost stays visible inside a narrow card instead of being pushed
       off the right edge by long descriptions. Description takes the slack. */
    /* Override the global table min-width:860px so the 5-column layout can shrink into
       a narrow/resized panel. Widths are percentages (not px) so Description keeps a
       usable share at any panel width; below min-width the wrapper scrolls instead of
       squeezing Description down to one character per line. */
    .tx-table {{ width: 100%; table-layout: fixed; min-width: 460px; }}
    .tx-table .nowrap {{ white-space: nowrap; }}
    .tx-table th:nth-child(1), .tx-table td:nth-child(1) {{ width: 17%; }}
    .tx-table th:nth-child(2), .tx-table td:nth-child(2) {{ width: 14%; }}
    .tx-table th:nth-child(4), .tx-table td:nth-child(4) {{ width: 17%; }}
    .tx-table th:nth-child(5), .tx-table td:nth-child(5) {{ width: 15%; }}
    .tx-table td:nth-child(3) {{ overflow-wrap: anywhere; }}
    /* Inline category editing on the account tabs needs a wider Category column. */
    .tx-table-editable {{ min-width: 520px; }}
    .tx-table-editable th:nth-child(4), .tx-table-editable td:nth-child(4) {{ width: 24%; }}
    .tx-table-editable select {{ width: 100%; max-width: 100%; font-size: 12px; padding: 3px 4px; border: 1px solid var(--line); border-radius: 6px; background: #fff; }}
    .override-note {{ font-size: 10px; text-transform: uppercase; letter-spacing: .04em; color: var(--green); font-weight: 700; margin-top: 2px; }}
    .rollup-detail .tx-table {{ font-size: 12px; }}
    .rollup-detail .tx-table td, .rollup-detail .tx-table th {{ padding: 5px 8px; }}
    /* Draggable split between the rollup and the transaction list. */
    .grid.split {{ display: grid; gap: 0 10px; align-items: start; }}
    .split-handle {{ position: relative; cursor: col-resize; align-self: stretch; min-height: 80px; touch-action: none; }}
    .split-handle::before {{ content: ''; position: absolute; top: 0; bottom: 0; left: 3px; width: 4px; background: var(--line); border-radius: 999px; opacity: .6; transition: background .1s, opacity .1s; }}
    .split-handle:hover::before, .split-handle.dragging::before, .split-handle:focus-visible::before {{ background: var(--green); opacity: 1; }}
    .split-handle:focus-visible {{ outline: none; }}
    .bar-fill {{ height: 100%; background: var(--accent-3); border-radius: inherit; }}
    .scroll-panel {{ max-height: 420px; overflow: auto; padding-right: 6px; scrollbar-width: thin; scrollbar-color: rgba(46,107,63,.45) transparent; }}
    .scroll-panel::-webkit-scrollbar {{ width: 8px; height: 8px; }}
    .scroll-panel::-webkit-scrollbar-thumb {{ background: rgba(46,107,63,.45); border-radius: 999px; }}
    .scroll-panel::-webkit-scrollbar-track {{ background: transparent; }}
    .filters {{ display: flex; flex-wrap: wrap; gap: 10px; margin: 0 0 14px; }}
    .table-wrap {{ overflow: auto; max-height: 620px; border: 1px solid var(--line); border-radius: var(--radius); }}
    .matrix-wrap {{ max-height: 680px; }}
    table {{ width: 100%; border-collapse: collapse; min-width: 860px; background: #fff; }}
    .matrix-table {{ min-width: max-content; }}
    th, td {{ padding: 10px 12px; border-bottom: 1px solid var(--line); text-align: left; font-size: 13px; vertical-align: top; }}
    th {{ position: sticky; top: 0; background: #f1f4ef; z-index: 1; color: #34413d; font-size: 12px; text-transform: uppercase; letter-spacing: .05em; }}
    .matrix-table th:first-child, .matrix-table td:first-child {{ position: sticky; left: 0; z-index: 2; background: #fff; min-width: 170px; }}
    .matrix-table th:first-child {{ z-index: 3; background: #f1f4ef; }}
    .matrix-table .right-break {{ border-left: 28px solid #fff; }}
    .matrix-table tr.subtotal-row td {{ border-top: 2px solid #1f2724; border-bottom: 2px solid #1f2724; font-weight: 850; background: #fff; }}
    .matrix-table tr.spacer-row td {{ height: 16px; padding: 0; border-bottom: 0; background: #fff; }}
    .matrix-table tr.margin-row td {{ font-style: italic; color: #384540; }}
    td.money, th.money {{ text-align: right; font-variant-numeric: tabular-nums; }}
    .range-controls {{ display: flex; flex-wrap: wrap; gap: 10px; align-items: end; margin-bottom: 14px; }}
    .range-controls label {{ display: grid; gap: 5px; color: var(--muted); font-size: 12px; font-weight: 800; text-transform: uppercase; letter-spacing: .05em; }}
    .tag {{ display: inline-flex; align-items: center; min-height: 24px; padding: 2px 8px; border-radius: 999px; background: #eef2f1; color: #31413d; font-size: 12px; font-weight: 700; white-space: nowrap; }}
    .tag.control-high {{ background: #dcebd8; color: #245435; }}
    .tag.control-medium {{ background: #efe7d7; color: #694b1e; }}
    .tag.control-low {{ background: #e4e9ec; color: #3c5360; }}
    .tag.forecast-status-overdue {{ background: #f6dddd; color: #7c2828; }}
    .tag.forecast-status-due {{ background: #efe7d7; color: #694b1e; }}
    .tag.forecast-status-charged {{ background: #dcebd8; color: #245435; }}
    tr.forecast-overdue td {{ background: #fff4f2; }}
    .chip-row {{ display: flex; flex-wrap: wrap; gap: 6px; margin-bottom: 8px; }}
    .control-card {{ display: grid; gap: 10px; }}
    .control-bar {{ height: 8px; background: #edf0eb; border-radius: 999px; overflow: hidden; }}
    .control-fill {{ height: 100%; border-radius: inherit; }}
    .control-fill.high {{ background: #2e6b3f; }}
    .control-fill.medium {{ background: #9e7132; }}
    .control-fill.low {{ background: #5b8aa6; }}
    .positive {{ color: var(--good); }}
    .negative {{ color: var(--danger); }}
    .insights {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 14px; }}
    .insight {{ border-left: 4px solid var(--accent-3); }}
    .insight.warning {{ border-left-color: var(--danger); }}
    .insight.opportunity {{ border-left-color: var(--accent); }}
    .account-title {{ display: flex; align-items: center; justify-content: space-between; gap: 12px; margin-bottom: 12px; }}
    .empty {{ color: var(--muted); padding: 22px; text-align: center; }}
    .legend {{ display: flex; flex-wrap: wrap; gap: 16px; margin-bottom: 8px; font-size: 12px; font-weight: 700; color: var(--ink); }}
    .legend-item {{ display: inline-flex; align-items: center; gap: 6px; }}
    .legend-dot {{ width: 10px; height: 10px; border-radius: 3px; display: inline-block; }}
    .delta-card .metric-note {{ display: flex; align-items: center; gap: 6px; font-weight: 700; }}
    .delta-up::before, .delta-down::before {{ content: ''; display: inline-block; width: 0; height: 0; border-left: 5px solid transparent; border-right: 5px solid transparent; }}
    .delta-up::before {{ border-bottom: 8px solid var(--good); }}
    .delta-down::before {{ border-top: 8px solid var(--danger); }}
    .sev {{ display: inline-block; padding: 1px 8px; border-radius: 999px; font-size: 11px; font-weight: 800; text-transform: uppercase; letter-spacing: .04em; background: #f3e3df; color: #7a2f1f; margin-bottom: 6px; }}
    .sev.high {{ background: #efcfc8; color: #6e1f14; }}
    @media (max-width: 1050px) {{
      header {{ grid-template-columns: 1fr; }}
      .toolbar {{ justify-content: flex-start; }}
      .kpis, .two, .three, .insights {{ grid-template-columns: 1fr 1fr; }}
      .grid.split {{ grid-template-columns: 1fr !important; gap: 16px; }}
      .split-handle {{ display: none; }}
    }}
    @media (max-width: 680px) {{
      .kpis, .two, .three, .insights {{ grid-template-columns: 1fr; }}
      .bar-row {{ grid-template-columns: 100px 1fr 76px; }}
      header {{ padding-top: 22px; }}
      nav button {{ min-width: 92px; }}
    }}
  </style>
</head>
<body>
  <header>
    <div>
      <h1>Financial Dashboard</h1>
      <div class="subtle" id="rangeLine"></div>
    </div>
    <div class="toolbar">
      <input id="globalSearch" type="search" placeholder="Search transactions">
      <select id="monthMode">
        <option value="all">All months</option>
        <option value="last6">Last 6 months</option>
        <option value="single">One month</option>
        <option value="range">Month range</option>
      </select>
      <select id="monthSingle"></select>
      <select id="monthStart"></select>
      <select id="monthEnd"></select>
      <button id="resetFilters">Reset</button>
    </div>
  </header>
  <nav id="tabs"></nav>
  <main id="app"></main>
  <div id="chartTip" role="tooltip" aria-hidden="true"></div>
  <script>
    const DATA = {data};
    const app = document.getElementById('app');
    const tabs = document.getElementById('tabs');
    const searchInput = document.getElementById('globalSearch');
    const monthMode = document.getElementById('monthMode');
    const monthSingle = document.getElementById('monthSingle');
    const monthStart = document.getElementById('monthStart');
    const monthEnd = document.getElementById('monthEnd');
    let activeTab = 'Overview';
    let forecastCategory = 'total';
    let analyticsCategory = null;
    let analyticsControlFilter = 'all';
    let monthlyRollupCat = null;  // Range Category Rollup: which bar is expanded
    let monthlySplit = 0.62;      // width fraction of the rollup card vs the tx list
    const overrideStorageKey = 'financial_dashboard_category_overrides_v1';
    const categoryLabels = {{
      alc: 'Alcohol',
      bar: 'Bars',
      bet: 'Betting',
      cash: 'Cash',
      check: 'Checks',
      clothing: 'Clothing',
      donations: 'Donations',
      event: 'Events',
      fee: 'Fees',
      food: 'Food',
      gift: 'Gifts',
      golf: 'Golf',
      grocery: 'Groceries',
      haircut: 'Haircuts',
      living: 'Living',
      metro: 'Metro',
      mom: 'Family transfers',
      payment: 'Card payments',
      rent: 'Rent',
      salary: 'Salary',
      service: 'Services',
      spotify: 'Spotify',
      stock: 'Investments',
      'additional savings': 'Additional Savings',
      tax: 'Taxes',
      travel: 'Travel',
      uber: 'Uber',
      utilities: 'Utilities',
      venmo: 'Venmo transfers',
      uncategorized: 'Uncategorized'
    }};
    const categoryChoices = Object.keys(categoryLabels).sort((a, b) => categoryLabels[a].localeCompare(categoryLabels[b]));
    const labelToCat = Object.fromEntries(Object.entries(categoryLabels).map(([key, label]) => [label, key]));
    labelToCat['mom'] = 'mom';
    const categoryControl = {{
      alc: 'high',
      bar: 'high',
      bet: 'high',
      cash: 'high',
      clothing: 'high',
      event: 'high',
      food: 'high',
      golf: 'high',
      haircut: 'high',
      living: 'high',
      metro: 'high',
      service: 'high',
      spotify: 'high',
      uber: 'high',
      donations: 'medium',
      gift: 'medium',
      grocery: 'medium',
      travel: 'medium',
      fee: 'low',
      rent: 'low',
      utilities: 'low'
    }};
    const controlLabels = {{high: 'High control', medium: 'Medium control', low: 'Low control', internal: 'Internal'}};
    let categoryOverrides = {{}};

    function loadCategoryOverrides() {{
      try {{
        categoryOverrides = JSON.parse(localStorage.getItem(overrideStorageKey) || '{{}}');
      }} catch {{
        categoryOverrides = {{}};
      }}
    }}

    function applyCategoryOverrides() {{
      DATA.transactions.forEach(tx => {{
        if (tx._origCategory === undefined) {{
          tx._origCategory = tx.category;
          tx._origCategoryLabel = tx.category_label;
        }}
        const category = categoryOverrides[tx.id];
        if (!category || !categoryLabels[category]) {{
          tx.category = tx._origCategory;
          tx.category_label = tx._origCategoryLabel;
          return;
        }}
        tx.category = category;
        tx.category_label = categoryLabels[category];
      }});
    }}

    // Overrides live in financials.sqlite when the page is served by serve_dashboard.py,
    // and in localStorage when it is opened as a plain file (a file:// page cannot write
    // to the database). The API is probed once at startup; everything else is identical.
    let overrideBackend = 'local';
    const overrideApi = 'api/overrides';

    async function initOverrideBackend() {{
      try {{
        const res = await fetch(overrideApi, {{cache: 'no-store'}});
        if (!res.ok) return;
        const data = await res.json();
        if (!data || typeof data !== 'object' || data.error) return;
        overrideBackend = 'server';
        categoryOverrides = data;
        applyCategoryOverrides();
        render();
      }} catch {{
        // No server (file:// or not running) -- stay on localStorage.
      }}
    }}

    function overrideStorageNote() {{
      return overrideBackend === 'server' ? 'Saved to financials.sqlite' : 'Stored in this browser';
    }}

    function setCategoryOverride(txId, category) {{
      if (!category || category === 'uncategorized') delete categoryOverrides[txId];
      else categoryOverrides[txId] = category;
      if (overrideBackend === 'server') {{
        fetch(overrideApi, {{
          method: 'POST',
          headers: {{'Content-Type': 'application/json'}},
          body: JSON.stringify({{tx_id: txId, category: category || ''}}),
        }}).then(res => {{
          if (!res.ok) console.error('Override not saved to the database', res.status);
        }}).catch(err => console.error('Override not saved to the database', err));
      }} else {{
        localStorage.setItem(overrideStorageKey, JSON.stringify(categoryOverrides));
      }}
      applyCategoryOverrides();
      render();
    }}

    const money = value => {{
      const n = Number(value || 0);
      const sign = n < 0 ? '-' : '';
      return sign + '$' + Math.abs(n).toLocaleString(undefined, {{minimumFractionDigits: 0, maximumFractionDigits: 0}});
    }};
    const exactMoney = value => {{
      const n = Number(value || 0);
      const sign = n < 0 ? '-' : '';
      return sign + '$' + Math.abs(n).toLocaleString(undefined, {{minimumFractionDigits: 2, maximumFractionDigits: 2}});
    }};
    const pct = value => (Number(value || 0) * 100).toFixed(0) + '%';
    const esc = value => String(value ?? '').replace(/[&<>"']/g, c => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[c]));
    const tipText = s => esc(esc(s));
    const niceMax = v => {{
      const n = Number(v) || 0;
      if (n <= 0) return 100;
      const exp = Math.floor(Math.log10(n));
      const base = Math.pow(10, exp);
      const frac = n / base;
      const nice = frac <= 1 ? 1 : frac <= 2 ? 2 : frac <= 5 ? 5 : 10;
      return nice * base;
    }};
    const moneyShort = value => {{
      const n = Number(value || 0);
      const sign = n < 0 ? '-' : '';
      const a = Math.abs(n);
      if (a >= 1000000) return sign + '$' + (a / 1000000).toFixed(a >= 10000000 ? 0 : 1).replace(/\\.0$/, '') + 'M';
      if (a >= 1000) return sign + '$' + (a / 1000).toFixed(a >= 10000 ? 0 : 1).replace(/\\.0$/, '') + 'k';
      return sign + '$' + Math.round(a);
    }};
    const incomeCategories = new Set(['salary']);
    const transferCategories = new Set(['payment','venmo','capone','amex','wells']);
    const savingsCategories = new Set(['stock','additional savings']);
    const internalCategories = new Set([...incomeCategories, ...transferCategories, ...savingsCategories]);
    const expenseCost = tx => internalCategories.has(tx.category) ? 0 : Number(tx.cost || 0);
    const grossExpenseCost = tx => Math.max(expenseCost(tx), 0);
    const budgetExpenseCost = tx => savingsCategories.has(tx.category) ? 0 : expenseCost(tx);
    const transferCost = tx => transferCategories.has(tx.category) ? Number(tx.cost || 0) : 0;
    const incomeValue = tx => incomeCategories.has(tx.category) ? Math.max(-Number(tx.cost || 0), 0) : 0;
    const savingsValue = tx => savingsCategories.has(tx.category) ? Math.max(Number(tx.cost || 0), 0) : 0;
    const controlLevel = category => {{
      if (internalCategories.has(category)) return 'internal';
      return categoryControl[category] || 'medium';
    }};
    const controlNote = level => ({{
      high: 'Direct behavior or merchant choice',
      medium: 'Plan, cap, or verify as one-off',
      low: 'Mostly fixed or required',
      internal: 'Excluded from budget spend'
    }}[level] || 'Plan, cap, or verify as one-off');
    const controlRecommendation = flag => {{
      const increase = Math.max(flag.latest - flag.base, 0);
      if (flag.control === 'high') return `Set next month near ${{money(flag.base)}} and inspect the largest merchants causing the ${{money(increase)}} jump.`;
      if (flag.control === 'medium') return `Decide whether the ${{money(increase)}} increase was planned; if not, set a category cap before it repeats.`;
      if (flag.control === 'low') return 'Verify the bill, timing, or duplicate charge before cutting flexible categories to offset it.';
      return 'Review the source transactions before treating this as spend.';
    }};
    const titleMerchant = value => String(value || '')
      .toLowerCase()
      .replace(/\\b\\w/g, c => c.toUpperCase())
      .replace(/\\bDc\\b/g, 'DC')
      .replace(/\\bAi\\b/g, 'AI')
      .replace(/\\bWsj\\b/g, 'WSJ');
    const merchantAliases = [
      [/\\b(walmart|wmt\\s*plus)\\b/i, 'Walmart'],
      [/\\b(amazon|amzn\\s*mktp|amzn\\.com)\\b/i, 'Amazon'],
      [/\\buber\\s*(eats|\\*eats)|help\\.uber\\.com/i, 'Uber Eats'],
      [/\\buber\\b/i, 'Uber'],
      [/\\blyft\\b/i, 'Lyft'],
      [/apple\\.com\\/bill/i, 'Apple'],
      [/\\bwalgreens\\b/i, 'Walgreens'],
      [/\\btarget(\\.com)?\\b/i, 'Target'],
      [/\\b(harris\\s+teeter)\\b/i, 'Harris Teeter'],
      [/\\b(safeway)\\b/i, 'Safeway'],
      [/\\b(wholefds|whole\\s*foods)\\b/i, 'Whole Foods'],
      [/\\btrader\\s+joe/i, "Trader Joe's"],
      [/\\b(comcast|xfinity)\\b/i, 'Comcast / Xfinity'],
      [/\\bdc\\s+residential\\s+water\\b/i, 'DC Residential Water'],
      [/\\bwashington\\s+gas\\b/i, 'Washington Gas'],
      [/\\bpaymentus\\b/i, 'Paymentus'],
      [/\\bstmat|st\\s*matthew/i, "St. Matthew's"],
      [/\\bkalshi\\b/i, 'Kalshi'],
      [/\\bspotify\\b/i, 'Spotify'],
      [/\\bmicrosoft\\b/i, 'Microsoft'],
      [/\\bclaude\\.ai|claude\\s*ai\\b/i, 'Claude AI'],
      [/\\b(wsjonline|wsj\\s+online|d\\s*j\\*wsj)\\b/i, 'Wall Street Journal'],
      [/\\bpanera\\b/i, 'Panera'],
      [/\\bstarbucks\\b/i, 'Starbucks'],
      [/\\bcircle\\s*k\\b/i, 'Circle K'],
      [/\\bpas?seto\\s+cleaning/i, 'Passeto Cleaning'],
      [/\\buniversal\\s+wine/i, 'Universal Wine & Liquor'],
      [/\\bfive\\s+iron\\s+golf/i, 'Five Iron Golf'],
      [/\\brailbird\\b/i, 'Railbird'],
      [/\\bsmartrip\\b/i, 'SmarTrip'],
      [/\\bmetro\\s+washington|^metro$/i, 'Metro']
    ];
    function merchantGroupLabel(tx) {{
      const raw = String((tx.counterparty || tx.description || '')).trim();
      if (!raw) return 'Unknown merchant';
      const alias = merchantAliases.find(([pattern]) => pattern.test(raw));
      if (alias) return alias[1];
      let cleaned = raw
        .replace(/^AplPay\\s+/i, '')
        .replace(/^TST\\*\\s*/i, '')
        .replace(/^SQ \\*/i, '')
        .replace(/^COT\\*/i, '')
        .replace(/^D J\\*/i, '')
        .replace(/^PAYPAL \\*/i, '')
        .replace(/\\b\\d{{3}}[- ]?\\d{{3}}[- ]?\\d{{4}}\\b/g, '')
        .replace(/\\b\\d{{1,2}}\\/\\d{{1,2}}\\b/g, '')
        .replace(/\\b\\d{{5,}}[A-Z]*\\b/gi, '')
        .replace(/#[A-Z0-9-]+/gi, '')
        .replace(/[|_/]+/g, ' ')
        .replace(/[.*]+/g, ' ')
        .replace(/\\.com\\b/ig, '')
        .replace(/\\b(AL|AK|AZ|AR|CA|CO|CT|DC|DE|FL|GA|HI|IA|ID|IL|IN|KS|KY|LA|MA|MD|ME|MI|MN|MO|MS|MT|NC|ND|NE|NH|NJ|NM|NV|NY|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT|VA|VT|WA|WI|WV|WY)\\b\\.?$/i, '')
        .replace(/\\b(WASHINGTON|BENTONVILLE|BROOKLYN PARK|SAN FRANCISCO|ROCKVILLE|DENVER|OAK RIDGE NORTH)\\b\\.?$/i, '')
        .replace(/\\s+/g, ' ')
        .trim();
      return cleaned ? titleMerchant(cleaned) : titleMerchant(raw);
    }}
    const allTransactionMonths = () => DATA.summary.months.map(r => r.month);
    const monthOptionsFor = (months, selected='') => months.map(month => `<option value="${{month}}" ${{month === selected ? 'selected' : ''}}>${{month}}</option>`).join('');
    const selectedGlobalMonths = (universe=allTransactionMonths()) => {{
      if (!universe.length) return [];
      const mode = monthMode.value;
      if (mode === 'last6') return universe.slice(-6);
      if (mode === 'single') return [monthSingle.value || universe.at(-1)].filter(month => universe.includes(month));
      if (mode === 'range') {{
        const start = monthStart.value || universe[0];
        const end = monthEnd.value || universe.at(-1);
        const lo = universe.indexOf(start);
        const hi = universe.indexOf(end);
        if (lo === -1 || hi === -1) return universe;
        return universe.slice(Math.min(lo, hi), Math.max(lo, hi) + 1);
      }}
      return universe;
    }};
    const selectedPeriodLabel = (months=selectedGlobalMonths()) => {{
      if (!months.length) return 'No months';
      if (months.length === allTransactionMonths().length) return 'All months';
      if (months.length === 1) return months[0];
      return `${{months[0]}} to ${{months.at(-1)}}`;
    }};
    const sourceCoverageLabel = months => {{
      const wb = new Set(DATA.workbook_monthly?.months || []);
      const workbookCount = months.filter(month => wb.has(month)).length;
      const actualCount = months.length - workbookCount;
      if (!actualCount) return 'workbook monthly model';
      if (!workbookCount) return 'actual transactions';
      return `${{workbookCount}} workbook months + ${{actualCount}} actual months`;
    }};
    const updateMonthControlVisibility = () => {{
      monthSingle.style.display = monthMode.value === 'single' ? '' : 'none';
      monthStart.style.display = monthMode.value === 'range' ? '' : 'none';
      monthEnd.style.display = monthMode.value === 'range' ? '' : 'none';
    }};
    const filtered = (account=null) => {{
      const query = searchInput.value.trim().toLowerCase();
      const monthUniverse = allTransactionMonths();
      const selectedMonths = new Set(selectedGlobalMonths(monthUniverse));
      return DATA.transactions.filter(tx => {{
        if (account && tx.account !== account) return false;
        if (monthUniverse.length && !selectedMonths.has(tx.month)) return false;
        if (!query) return true;
        return [tx.account, tx.date, tx.description, tx.category_label, tx.native_category, tx.counterparty].join(' ').toLowerCase().includes(query);
      }});
    }};

    function kpi(label, value, note='') {{
      return `<div class="card"><div class="metric-label">${{esc(label)}}</div><div class="metric-value">${{value}}</div><div class="metric-note">${{esc(note)}}</div></div>`;
    }}

    function deltaCard(label, cur, delta, pctChg, isNew) {{
      const up = delta >= 0;
      const sign = up ? '+' : '-';
      const word = up ? 'higher' : 'lower';
      const arrow = up ? 'delta-up' : 'delta-down';
      const change = isNew
        ? 'New this month'
        : `${{sign}}${{money(Math.abs(delta))}} ${{word}}${{pctChg == null ? '' : ` (${{pct(Math.abs(pctChg))}})`}}`;
      const noteClass = isNew ? '' : arrow;
      return `<div class="card delta-card"><div class="metric-label">${{esc(label)}}</div><div class="metric-value">${{money(cur)}}</div><div class="metric-note"><span class="${{noteClass}}"></span>${{esc(change)}}</div></div>`;
    }}

    function lineChart(rows, series=[{{key: 'expenses', color: '#2e6b3f', label: 'Spend'}}]) {{
      const w = 920, h = 280, padL = 64, padR = 24, padTop = 16, padBot = 34;
      const single = series.length === 1;
      const dataMax = Math.max(...rows.flatMap(r => series.map(s => Number(r[s.key] || 0))), 0);
      const max = niceMax(dataMax || 100);
      const xAt = i => padL + (rows.length <= 1 ? 0 : i * (w - padL - padR) / (rows.length - 1));
      const yAt = v => h - padBot - (Number(v || 0) / max) * (h - padBot - padTop);
      const ticks = 5;
      const gridlines = Array.from({{length: ticks}}, (_, i) => {{
        const val = max * i / (ticks - 1);
        const y = yAt(val);
        return `<line x1="${{padL}}" y1="${{y.toFixed(1)}}" x2="${{w - padR}}" y2="${{y.toFixed(1)}}" stroke="var(--line)" stroke-width="1"></line>
          <text x="${{padL - 8}}" y="${{(y + 4).toFixed(1)}}" text-anchor="end" font-size="11" fill="#64706b">${{moneyShort(val)}}</text>`;
      }}).join('');
      const baseline = `<line x1="${{padL}}" y1="${{h - padBot}}" x2="${{w - padR}}" y2="${{h - padBot}}" stroke="var(--line)" stroke-width="1"></line>`;
      const lines = series.map(s => {{
        const points = rows.map((r, i) => [xAt(i), yAt(r[s.key]), r]);
        const path = points.map((p, i) => `${{i ? 'L' : 'M'}}${{p[0].toFixed(1)}} ${{p[1].toFixed(1)}}`).join(' ');
        const area = single
          ? `<path d="${{path}} L ${{points.at(-1)?.[0] || padL}} ${{h-padBot}} L ${{padL}} ${{h-padBot}} Z" fill="#dde9cd"></path>`
          : '';
        return `${{area}}<path d="${{path}}" fill="none" stroke="${{s.color}}" stroke-width="4" stroke-linecap="round"></path>`;
      }}).join('');
      const dots = rows.map((r, i) => {{
        const x = xAt(i);
        const tip = `${{tipText(r.month)}}<br>` + series.map(s => `${{tipText(s.label)}}: ${{money(r[s.key])}}`).join('<br>');
        return series.map(s => {{
          const y = yAt(r[s.key]);
          return `<g class="ptg"><circle class="pt" cx="${{x.toFixed(1)}}" cy="${{y.toFixed(1)}}" r="4" fill="${{s.color}}"></circle><circle cx="${{x.toFixed(1)}}" cy="${{y.toFixed(1)}}" r="11" fill="transparent" data-tip="${{tip}}"></circle></g>`;
        }}).join('');
      }}).join('');
      const axisLabel = month => {{
        const [year, monthNum] = month.split('-');
        return `${{Number(monthNum)}}/${{year.slice(2)}}`;
      }};
      const labels = rows.map((r, i) => `<text x="${{xAt(i)}}" y="${{h-12}}" text-anchor="middle" font-size="10" fill="#64706b">${{axisLabel(r.month)}}</text>`).join('');
      const legend = series.length > 1
        ? `<div class="legend">${{series.map(s => `<span class="legend-item"><span class="legend-dot" style="background:${{s.color}}"></span>${{esc(s.label)}}</span>`).join('')}}</div>`
        : '';
      return `${{legend}}<svg class="chart" viewBox="0 0 ${{w}} ${{h}}" role="img" aria-label="Monthly trend with value axis">
        ${{gridlines}}
        ${{baseline}}
        ${{lines}}
        ${{dots}}
        ${{labels}}
      </svg>`;
    }}

    function forecastChart(rows) {{
      const series = [
        {{key: 'actual', color: 'var(--accent)', label: 'Actual', width: 4}},
        {{key: 'committed', color: 'var(--accent-3)', label: 'Committed path', width: 3, dash: '7 5'}},
        {{key: 'low', color: 'var(--muted)', label: 'Low', width: 3}},
        {{key: 'medium', color: 'var(--gold)', label: 'Medium', width: 3}},
        {{key: 'high', color: 'var(--danger)', label: 'High', width: 3}},
      ];
      const w = 920, h = 280, padL = 64, padR = 24, padTop = 16, padBot = 34;
      const values = rows.flatMap(row => series.map(item => row[item.key]).filter(value => value != null).map(Number));
      let min = Math.min(0, ...values), max = Math.max(0, ...values);
      if (min === max) max = min + 100;
      const span = max - min;
      const xAt = i => padL + (rows.length <= 1 ? 0 : i * (w - padL - padR) / (rows.length - 1));
      const yAt = value => padTop + (max - Number(value)) / span * (h - padBot - padTop);
      const gridlines = Array.from({{length: 5}}, (_, index) => {{
        const value = min + span * index / 4;
        const y = yAt(value);
        return `<line x1="${{padL}}" y1="${{y.toFixed(1)}}" x2="${{w - padR}}" y2="${{y.toFixed(1)}}" stroke="var(--line)" stroke-width="1"></line>
          <text x="${{padL - 8}}" y="${{(y + 4).toFixed(1)}}" text-anchor="end" font-size="11" fill="var(--muted)">${{moneyShort(value)}}</text>`;
      }}).join('');
      const lines = series.map(item => {{
        const points = rows
          .map((row, index) => row[item.key] == null ? null : [xAt(index), yAt(row[item.key])])
          .filter(Boolean);
        if (!points.length) return '';
        const path = points.map((point, index) => `${{index ? 'L' : 'M'}}${{point[0].toFixed(1)}} ${{point[1].toFixed(1)}}`).join(' ');
        return `<path d="${{path}}" fill="none" stroke="${{item.color}}" stroke-width="${{item.width}}" stroke-dasharray="${{item.dash || ''}}" stroke-linecap="round" stroke-linejoin="round"></path>`;
      }}).join('');
      const dots = rows.map((row, index) => {{
        const available = series.filter(item => row[item.key] != null);
        if (!available.length) return '';
        const tip = `Day ${{row.day}}<br>` + available.map(item => `${{tipText(item.label)}}: ${{money(row[item.key])}}`).join('<br>');
        return available.map(item => `<g class="ptg"><circle class="pt" cx="${{xAt(index).toFixed(1)}}" cy="${{yAt(row[item.key]).toFixed(1)}}" r="3" fill="${{item.color}}"></circle><circle cx="${{xAt(index).toFixed(1)}}" cy="${{yAt(row[item.key]).toFixed(1)}}" r="10" fill="transparent" data-tip="${{tip}}"></circle></g>`).join('');
      }}).join('');
      const today = Number(DATA.forecast?.day_of_month || 0);
      const labels = rows.map((row, index) =>
        (row.day === 1 || row.day === today || row.day === rows.at(-1)?.day || row.day % 5 === 0)
          ? `<text x="${{xAt(index)}}" y="${{h - 12}}" text-anchor="middle" font-size="10" fill="var(--muted)">${{row.day}}</text>`
          : ''
      ).join('');
      const legend = `<div class="legend">${{series.map(item => `<span class="legend-item"><span class="legend-dot" style="background:${{item.color}}"></span>${{esc(item.label)}}</span>`).join('')}}</div>`;
      return `${{legend}}<svg class="chart" viewBox="0 0 ${{w}} ${{h}}" role="img" aria-label="Cumulative spend forecast by day of month">
        ${{gridlines}}${{lines}}${{dots}}${{labels}}
      </svg>`;
    }}

    function bars(rows, valueKey='expense', limit=10, colorBy='category') {{
      const top = rows.filter(r => Number(r[valueKey]) > 0).slice(0, limit);
      const max = Math.max(...top.map(r => Number(r[valueKey])), 1);
      const total = top.reduce((sum, r) => sum + Number(r[valueKey]), 0) || 1;
      return `<div class="bars">${{top.map(r => {{
        const cat = colorBy === 'category' ? (r.category || labelToCat[r.label]) : null;
        const fill = colorBy === 'category' ? `;background:var(--cat-${{cat}}, var(--cat-default))` : '';
        const label = r.label || r.account || r.category;
        const share = Number(r[valueKey]) / total * 100;
        const tip = `${{tipText(label)}}<br>${{money(r[valueKey])}}<br>${{share.toFixed(1)}}% of shown`
          + (r.count != null ? `<br>${{r.count}} transactions` : '')
          + (r.variantCount != null && r.variantCount > 1 ? `<br>${{r.variantCount}} matched descriptions` : '');
        return `<div class="bar-row" data-tip="${{tip}}"><div>${{esc(label)}}</div><div class="bar-track"><div class="bar-fill" style="width:${{Math.max(2, Number(r[valueKey]) / max * 100)}}%${{fill}}"></div></div><div class="money">${{money(r[valueKey])}}</div></div>`;
      }}).join('')}}</div>`;
    }}

    // Range Category Rollup bars, but each row expands to the transactions behind it.
    // Mirrors bars() for layout; the workbook's row labels are lowercase for some
    // categories ('travel', 'mom'), so resolve them the same way workbookMonthlyRows does.
    function rollupCatFor(row) {{
      return row.category || labelToCat[row.label] || String(row.label || '').toLowerCase();
    }}

    function rollupBars(rows, months) {{
      // Sorted by net cost, highest spend first, so the order always reflects the active
      // filter. Net-negative categories (credits/reimbursements such as mom) are kept and
      // fall to the bottom; their bar is drawn from |total| and marked as a credit.
      const list = rows
        .filter(r => Math.round(Number(r.total) * 100) !== 0)
        .sort((a, b) => Number(b.total) - Number(a.total));
      if (!list.length) return '<div class="empty">No categories with activity in the selected period.</div>';
      const maxAbs = Math.max(...list.map(r => Math.abs(Number(r.total))), 1);
      const spendTotal = list.reduce((sum, r) => sum + Math.max(Number(r.total), 0), 0) || 1;
      return `<div class="bars">${{list.map(r => {{
        const cat = rollupCatFor(r);
        const value = Number(r.total);
        const open = monthlyRollupCat === cat;
        const credit = value < 0;
        const share = value / spendTotal * 100;
        const tip = `${{tipText(r.label)}}<br>${{money(value)}}<br>`
          + (credit ? 'Credit / reimbursement — offsets spend' : `${{share.toFixed(1)}}% of spend shown`)
          + '<br>Click to see the transactions';
        return `<div class="bar-row rollup-row" data-rollup-cat="${{esc(cat)}}" data-tip="${{tip}}" role="button" tabindex="0" aria-expanded="${{open}}">
            <div><span class="rollup-caret">${{open ? '&#9662;' : '&#9656;'}}</span>${{esc(r.label)}}</div>
            <div class="bar-track"><div class="bar-fill${{credit ? ' credit' : ''}}" style="width:${{Math.max(2, Math.abs(value) / maxAbs * 100)}}%;background:var(--cat-${{cat}}, var(--cat-default))"></div></div>
            <div class="money${{credit ? ' negative' : ''}}">${{money(value)}}</div>
          </div>${{open ? `<div class="rollup-detail">${{rollupDetail(cat, months)}}</div>` : ''}}`;
      }}).join('')}}</div>`;
    }}

    // Drag (or arrow-key) the divider to trade width between the rollup and the tx list.
    // The fraction lives in monthlySplit so it survives the re-render on every expand.
    function wireSplitHandle() {{
      const handle = document.getElementById('monthlySplitHandle');
      const grid = document.getElementById('monthlySplitGrid');
      if (!handle || !grid) return;
      const apply = fraction => {{
        monthlySplit = Math.min(0.82, Math.max(0.18, fraction));
        grid.style.gridTemplateColumns = `${{(monthlySplit * 100).toFixed(2)}}% 10px minmax(0, 1fr)`;
      }};
      const onMove = event => {{
        const rect = grid.getBoundingClientRect();
        if (!rect.width) return;
        const x = (event.touches ? event.touches[0].clientX : event.clientX) - rect.left;
        apply(x / rect.width);
      }};
      const stop = () => {{
        handle.classList.remove('dragging');
        document.body.style.userSelect = '';
        document.removeEventListener('mousemove', onMove);
        document.removeEventListener('mouseup', stop);
        document.removeEventListener('touchmove', onMove);
        document.removeEventListener('touchend', stop);
      }};
      const start = event => {{
        event.preventDefault();
        handle.classList.add('dragging');
        document.body.style.userSelect = 'none';
        document.addEventListener('mousemove', onMove);
        document.addEventListener('mouseup', stop);
        document.addEventListener('touchmove', onMove, {{passive: false}});
        document.addEventListener('touchend', stop);
      }};
      handle.addEventListener('mousedown', start);
      handle.addEventListener('touchstart', start, {{passive: false}});
      handle.addEventListener('dblclick', () => apply(0.62));
      handle.addEventListener('keydown', event => {{
        if (event.key === 'ArrowLeft') {{ event.preventDefault(); apply(monthlySplit - 0.02); }}
        if (event.key === 'ArrowRight') {{ event.preventDefault(); apply(monthlySplit + 0.02); }}
      }});
    }}

    function rollupDetail(cat, months) {{
      const monthSet = new Set(months);
      const rows = DATA.transactions
        .filter(tx => tx.category === cat && monthSet.has(tx.month))
        .sort((a, b) => (a.date === b.date ? 0 : (a.date < b.date ? 1 : -1)));
      if (!rows.length) return '<div class="empty">No transactions for this category in the selected period.</div>';
      const sum = rows.reduce((s, tx) => s + Number(tx.cost || 0), 0);
      return `<div class="subtle" style="margin:8px 0 6px">${{rows.length}} transaction${{rows.length === 1 ? '' : 's'}}`
        + ` &middot; ${{money(sum)}} &middot; ${{esc(selectedPeriodLabel(months))}}</div>${{txTable(rows)}}`;
    }}

    // Date / Account / Description / Category / Cost. Native, Status and Amount were
    // dropped: Amount duplicates Cost up to sign, and the other two are rarely useful
    // next to the description. Fixed layout keeps Cost on screen in narrow cards.
    // editable=true swaps the Category tag for the same <select> the Review tab uses, so
    // a category can be corrected from whichever account tab the row showed up on.
    function txTable(rows, editable = false) {{
      if (!rows.length) return '<div class="empty">No transactions match the current filters.</div>';
      return `<div class="table-wrap"><table class="tx-table${{editable ? ' tx-table-editable' : ''}}">
        <thead><tr><th>Date</th><th>Account</th><th>Description</th><th>Category</th><th class="money">Cost</th></tr></thead>
        <tbody>${{rows.map(tx => `<tr>
          <td class="nowrap">${{esc(tx.date)}}</td>
          <td>${{esc(tx.account)}}</td>
          <td>${{esc(tx.description)}}${{tx.counterparty ? `<div class="subtle">${{esc(tx.counterparty)}}</div>` : ''}}</td>
          <td>${{editable ? categorySelect(tx) : `<span class="tag">${{esc(tx.category_label)}}</span>`}}${{categoryOverrides[tx.id] ? '<div class="subtle override-note">edited</div>' : ''}}</td>
          <td class="money ${{Number(tx.cost) >= 0 ? '' : 'negative'}}">${{exactMoney(tx.cost)}}</td>
        </tr>`).join('')}}</tbody>
      </table></div>`;
    }}

    // Shared by the Review tab and every account tab.
    function wireCategorySelects(root) {{
      (root || document).querySelectorAll('[data-review-category]').forEach(select => {{
        select.addEventListener('change', event => setCategoryOverride(event.target.dataset.reviewCategory, event.target.value));
      }});
    }}

    function categorySelect(tx) {{
      return `<select data-review-category="${{esc(tx.id)}}" aria-label="Select category">
        ${{categoryChoices.map(category => `<option value="${{category}}" ${{tx.category === category ? 'selected' : ''}}>${{esc(categoryLabels[category])}}</option>`).join('')}}
      </select>`;
    }}

    function uncertainTransactions() {{
      const weakCategories = new Set(['uncategorized', 'merchandise']);
      return filtered().filter(tx => {{
        // Already decided by hand -- stop asking.
        if (categoryOverrides[tx.id]) return false;
        if (weakCategories.has(tx.category)) return true;
        // Flagged by the builder as a food/bar toss-up.
        return Number(tx.needs_review) === 1;
      }});
    }}

    function reviewReason(tx) {{
      if (tx.category === 'uncategorized') return 'No rule matched';
      if (tx.category === 'merchandise') return 'Generic label';
      return 'Food or bar?';
    }}

    function reviewTable(rows) {{
      if (!rows.length) return '<div class="empty">No uncertain transactions match the current filters.</div>';
      return `<div class="table-wrap"><table>
        <thead><tr><th>Date</th><th>Account</th><th>Description</th><th>Why</th><th>Current</th><th>Set Category</th><th class="money">Cost</th><th>Source</th></tr></thead>
        <tbody>${{rows.map(tx => `<tr>
          <td>${{esc(tx.date)}}</td>
          <td>${{esc(tx.account)}}</td>
          <td>${{esc(tx.description)}}${{tx.counterparty ? `<div class="subtle">${{esc(tx.counterparty)}}</div>` : ''}}</td>
          <td class="subtle">${{esc(reviewReason(tx))}}</td>
          <td><span class="tag">${{esc(tx.category_label)}}</span></td>
          <td>${{categorySelect(tx)}}</td>
          <td class="money ${{Number(tx.cost) < 0 ? 'negative' : ''}}">${{exactMoney(tx.cost)}}</td>
          <td>${{esc(tx.source)}}</td>
        </tr>`).join('')}}</tbody>
      </table></div>`;
    }}

    function monthlySourceMonths() {{
      const union = new Set([...(DATA.workbook_monthly?.months || []), ...allTransactionMonths()]);
      return [...union].sort();
    }}

    function monthOptions(selected='') {{
      return monthlySourceMonths().map(month => `<option value="${{month}}" ${{month === selected ? 'selected' : ''}}>${{month}}</option>`).join('');
    }}

    function selectedMonthlyMonths() {{
      const source = new Set(monthlySourceMonths());
      return selectedGlobalMonths().filter(month => source.has(month));
    }}

    function monthEndLabel(month) {{
      const [year, monthNum] = month.split('-').map(Number);
      const d = new Date(year, monthNum, 0);
      return `${{d.getMonth() + 1}}/${{d.getDate()}}/${{String(d.getFullYear()).slice(2)}}`;
    }}

    function accounting(value, decimals=2) {{
      const n = Number(value || 0);
      if (Math.abs(n) < 0.005) return '-';
      const body = Math.abs(n).toLocaleString(undefined, {{minimumFractionDigits: decimals, maximumFractionDigits: decimals}});
      return n < 0 ? `(${{body}})` : body;
    }}

    function percentAccounting(value) {{
      const n = Number(value || 0);
      if (!Number.isFinite(n) || Math.abs(n) < 0.0005) return '-';
      const body = Math.abs(n * 100).toFixed(1) + '%';
      return n < 0 ? `(${{body}})` : body;
    }}

    function monthSum(month, predicate, valueFn) {{
      return DATA.transactions.reduce((sum, tx) => sum + (tx.month === month && predicate(tx) ? valueFn(tx) : 0), 0);
    }}

    function retirement401k(month) {{
      return month < '2025-07' ? 779.17 : 830.50;
    }}

    function iraContribution(month) {{
      return month < '2025-07' ? 141.67 : 151.00;
    }}

    function monthlyRows(months) {{
      const categoryOrder = ['food','travel','service','grocery','bar','utilities','living','uber','clothing','gift','fee','donations','golf','alc','haircut','bet','cash','metro','spotify','mom'];
      const labels = new Map(DATA.transactions.map(tx => [tx.category, tx.category_label]));
      const included = new Set(categoryOrder);
      DATA.transactions.forEach(tx => {{
        if (tx.category === 'rent' || tx.category === 'salary' || tx.category === 'check' || tx.category === 'stock') return;
        if (transferCategories.has(tx.category) || savingsCategories.has(tx.category)) return;
        if (tx.category !== 'uncategorized') included.add(tx.category);
      }});
      return [...included].map(category => {{
        const values = Object.fromEntries(months.map(month => [month, monthSum(month, tx => tx.category === category, tx => Number(tx.cost || 0))]));
        const count = DATA.transactions.filter(tx => months.includes(tx.month) && tx.category === category).length;
        return {{
          category,
          label: category === 'mom' ? 'mom' : labels.get(category) || category,
          group: 'Spending',
          count,
          values,
        }};
      }}).map(row => {{
        row.total = months.reduce((sum, m) => sum + Number(row.values[m] || 0), 0);
        row.average = row.total / Math.max(months.length, 1);
        return row;
      }}).filter(row => Math.abs(row.total) > 0.005 || categoryOrder.includes(row.category)).sort((a, b) => {{
        const ai = categoryOrder.indexOf(a.category);
        const bi = categoryOrder.indexOf(b.category);
        if (ai !== -1 || bi !== -1) return (ai === -1 ? 999 : ai) - (bi === -1 ? 999 : bi);
        return Math.abs(b.total) - Math.abs(a.total);
      }});
    }}

    function monthlySummary(months) {{
      const rows = workbookMonthlyRows(months);
      const rowTotal = label => rows.find(row => row.label === label)?.total || 0;
      return {{
        income: rowTotal('Salary'),
        expenses: rowTotal('Total'),
        savings: rowTotal('Total Inv Savings'),
        netIncome: rowTotal('Net Income'),
        totalSavings: rowTotal('Total Savings'),
        transactions: DATA.transactions.filter(tx => months.includes(tx.month)).length,
      }};
    }}

    function workbookMonthlyRows(months) {{
      const sourceRows = DATA.workbook_monthly?.rows || [];
      const wbMonths = new Set(DATA.workbook_monthly?.months || []);
      const spendingLabels = ['Food','travel','service','grocery','Bar','Utilities','Living','uber','clothing','Gift','fee','Donations','Golf','Alc','Haircut','Bet','cash','Metro','Spotify','mom'];
      const derived = {{}};
      months.filter(month => !wbMonths.has(month)).forEach(month => {{
        const vals = {{}};
        let total = 0;
        spendingLabels.forEach(label => {{
          const cat = labelToCat[label] || label.toLowerCase();
          const v = monthSum(month, tx => tx.category === cat, tx => Number(tx.cost || 0));
          vals[label] = v;
          total += v;
        }});
        const rentVal = -monthSum(month, tx => tx.category === 'rent', tx => Number(tx.cost || 0));
        const salaryVal = -monthSum(month, tx => tx.category === 'salary', tx => Number(tx.cost || 0));
        const k401 = retirement401k(month);
        const ira = iraContribution(month);
        const additionalSavings = 0;
        const netIncome = salaryVal - total + rentVal;
        const totalInvSavings = k401 + ira + additionalSavings;
        vals['Rent'] = rentVal;
        vals['Salary'] = salaryVal;
        vals['401k'] = k401;
        vals['IRA'] = ira;
        vals['Additional Savings'] = additionalSavings;
        vals['Total'] = total;
        vals['Net Income'] = netIncome;
        vals['Margin'] = salaryVal ? netIncome / salaryVal : 0;
        vals['Total Inv Savings'] = totalInvSavings;
        vals['Total Savings'] = netIncome + totalInvSavings;
        derived[month] = vals;
      }});
      const rows = sourceRows.map(row => {{
        if (row.style === 'spacer-row') return {{...row, total: null, average: null, use: null}};
        const values = Object.fromEntries(months.map(month => [month, wbMonths.has(month) ? Number(row.values?.[month] || 0) : (derived[month]?.[row.label] ?? 0)]));
        const total = row.kind === 'percent' || row.kind === 'integer' && !row.label ? null : months.reduce((sum, month) => sum + Number(values[month] || 0), 0);
        const average = total === null ? null : total / Math.max(months.length, 1);
        return {{...row, values, total, average, use: total}};
      }});
      const salary = rows.find(row => row.label === 'Salary');
      const netIncome = rows.find(row => row.label === 'Net Income');
      const margin = rows.find(row => row.label === 'Margin');
      if (margin) {{
        const marginValues = margin.values || {{}};
        margin.total = salary?.total ? (netIncome?.total || 0) / salary.total : 0;
        margin.average = months.reduce((sum, month) => sum + Number(marginValues[month] || 0), 0) / Math.max(months.length, 1);
        margin.use = margin.total;
      }}
      return rows;
    }}

    function workbookRowValue(label, month) {{
      const row = DATA.workbook_monthly?.rows?.find(item => item.label === label);
      return Number(row?.values?.[month] || 0);
    }}

    function budgetSpendForMonth(month) {{
      if ((DATA.workbook_monthly?.months || []).includes(month)) {{
        return Math.max(workbookRowValue('Total', month) + Math.abs(workbookRowValue('Rent', month)), 0);
      }}
      const rows = DATA.transactions.filter(tx => tx.month === month);
      const categorySpend = rows.reduce((sum, tx) => {{
        if (tx.category === 'rent') return sum;
        return sum + budgetExpenseCost(tx);
      }}, 0);
      const rent = rows.reduce((sum, tx) => sum + (tx.category === 'rent' ? Math.abs(Number(tx.cost || 0)) : 0), 0);
      return Math.max(categorySpend + rent, 0);
    }}

    function budgetSpendPartsForMonth(month) {{
      if ((DATA.workbook_monthly?.months || []).includes(month)) {{
        const spend = Math.max(workbookRowValue('Total', month), 0);
        const rent = Math.abs(workbookRowValue('Rent', month));
        return {{spend, rent, total: spend + rent}};
      }}
      const rows = DATA.transactions.filter(tx => tx.month === month);
      const spend = Math.max(rows.reduce((sum, tx) => {{
        if (tx.category === 'rent') return sum;
        return sum + budgetExpenseCost(tx);
      }}, 0), 0);
      const rent = rows.reduce((sum, tx) => sum + (tx.category === 'rent' ? Math.abs(Number(tx.cost || 0)) : 0), 0);
      return {{spend, rent, total: spend + rent}};
    }}

    function monthlyMatrixTable() {{
      const months = selectedMonthlyMonths();
      const rows = workbookMonthlyRows(months);
      const wbMonths = new Set(DATA.workbook_monthly?.months || []);
      const monthHeader = month => wbMonths.has(month)
        ? `<th class="money">${{monthEndLabel(month)}}</th>`
        : `<th class="money" title="actual, from transactions">${{monthEndLabel(month)}}<span class="subtle">*</span></th>`;
      const hasDerived = months.some(month => !wbMonths.has(month));
      const renderValue = (row, value) => row.kind === 'percent' ? percentAccounting(value) : row.kind === 'integer' ? accounting(value, 0) : accounting(value);
      const renderRight = (row, key, extraClass='') => `<td class="money right-break ${{extraClass}}">${{renderValue(row, row[key] ?? 0)}}</td>`;
      const renderRow = (row, className='') => `<tr class="${{className}}">
        <td><span class="swatch" style="background:var(--cat-${{labelToCat[row.label]}}, var(--cat-default))"></span>${{esc(row.label)}}</td>
        ${{months.map(month => `<td class="money">${{renderValue(row, row.values[month] || 0)}}</td>`).join('')}}
        ${{renderRight(row, 'total')}}
        <td class="money">${{renderValue(row, row.average || 0)}}</td>
        <td class="money right-break">${{renderValue(row, row.use ?? row.total ?? 0)}}</td>
      </tr>`;
      return `<div class="table-wrap matrix-wrap"><table class="matrix-table">
        <thead><tr><th>Category</th>${{months.map(monthHeader).join('')}}<th class="money right-break">Range Total</th><th class="money">Avg</th><th class="money right-break">Use</th></tr></thead>
        <tbody>
          ${{rows.map(row => row.style === 'spacer-row' ? `<tr class="spacer-row"><td colspan="${{months.length + 4}}"></td></tr>` : renderRow(row, row.style || '')).join('')}}
        </tbody>
      </table></div>${{hasDerived ? '<div class="subtle">* computed from transactions (not in the budget workbook)</div>' : ''}}`;
    }}

    function summarizeTransactions(rows) {{
      const income = rows.reduce((sum, tx) => sum + incomeValue(tx), 0);
      const months = [...new Set(rows.map(tx => tx.month))].sort();
      const expenses = months.reduce((sum, month) => sum + budgetSpendForMonth(month), 0);
      const savings = rows.reduce((sum, tx) => sum + savingsValue(tx), 0);
      const rent = rows.reduce((sum, tx) => sum + (tx.category === 'rent' ? Number(tx.cost || 0) : 0), 0);
      return {{
        income,
        expenses,
        savings,
        rent,
        transactions: rows.length,
        accounts: new Set(rows.map(tx => tx.account)).size,
        net: income - expenses,
      }};
    }}

    function categoryRollup(rows) {{
      return Object.values(rows.reduce((acc, tx) => {{
        if (!acc[tx.category]) acc[tx.category] = {{category: tx.category, label: tx.category_label, expense: 0, count: 0}};
        acc[tx.category].expense += budgetExpenseCost(tx);
        acc[tx.category].count += 1;
        return acc;
      }}, {{}})).map(row => ({{...row, expense: Math.max(row.expense, 0)}})).sort((a,b) => b.expense - a.expense);
    }}

    function accountRollup(rows) {{
      return Object.values(rows.reduce((acc, tx) => {{
        if (!acc[tx.account]) acc[tx.account] = {{label: tx.account, expense: 0, count: 0}};
        acc[tx.account].expense += budgetExpenseCost(tx);
        acc[tx.account].count += 1;
        return acc;
      }}, {{}})).map(row => ({{...row, expense: Math.max(row.expense, 0)}})).sort((a,b) => b.expense - a.expense);
    }}

    function monthlyExpenseRows(months, rows) {{
      return months.map(month => {{
        const parts = budgetSpendPartsForMonth(month);
        return {{month, expenses: parts.total, spend: parts.spend, rent: parts.rent}};
      }});
    }}

    function trendValueStrip(rows) {{
      const shown = rows.slice(-6);
      const max = Math.max(...shown.map(r => r.expenses), 1);
      return `<div class="bars" style="margin-top:8px">${{shown.map(row => `<div class="bar-row"><div>${{esc(row.month)}}<div class="subtle">Spend ${{money(row.spend)}} · Rent ${{money(row.rent)}}</div></div><div class="bar-track"><div class="bar-fill" style="width:${{Math.max(2, row.expenses / max * 100)}}%"></div></div><div class="money">${{money(row.expenses)}}</div></div>`).join('')}}</div>`;
    }}

    function renderOverview() {{
      const months = selectedGlobalMonths();
      const rows = filtered();
      const s = summarizeTransactions(rows);
      const avg = s.expenses / Math.max(months.length, 1);
      const trendRows = monthlyExpenseRows(months, rows);
      app.innerHTML = `
        <section class="grid kpis">
          ${{kpi('Period', selectedPeriodLabel(months), `${{months.length}} months · ${{sourceCoverageLabel(months)}}`)}}
          ${{kpi('Loaded transactions', s.transactions.toLocaleString(), `${{s.accounts}} accounts`)}}
          ${{kpi('Income', money(s.income), 'Salary')}}
          ${{kpi('Expenses', money(s.expenses), `Avg ${{money(avg)}} per selected month · includes actual months`)}}
          ${{kpi('Rent', money(s.rent), (s.expenses < 1 ? '-' : pct(s.rent / s.expenses)) + ' of spend')}}
          ${{kpi('Investments', money(s.savings), 'Running investment total')}}
          ${{kpi('Budget Net', money(s.net), 'Income minus expenses')}}
        </section>
        <section class="grid two" style="margin-top:16px">
          <div class="card"><h2>Monthly Spend Trend</h2>${{lineChart(trendRows, [
            {{key: 'expenses', color: '#2e6b3f', label: 'Total'}},
            {{key: 'spend', color: '#345995', label: 'Spend'}}
          ])}}${{trendValueStrip(trendRows)}}</div>
          <div class="card"><h2>Top Categories</h2>${{bars(categoryRollup(rows), 'expense', 12)}}</div>
        </section>
        <section class="grid two" style="margin-top:16px">
          <div class="card"><h2>Accounts</h2>${{bars(accountRollup(rows), 'expense', 8, null)}}</div>
          <div class="card"><h2>Recent Transactions</h2>${{txTable(filtered().slice(0, 20))}}</div>
        </section>`;
    }}

    function renderForecast() {{
      const forecast = DATA.forecast || {{total: {{}}, committed: {{items: []}}, categories: [], cumulative: []}};
      const categories = forecast.categories || [];
      const committed = forecast.committed || {{remaining: 0, charged_so_far: 0, items: []}};
      const calendar = forecast.calendar || {{weekdays_remaining: 0, weekend_days_remaining: 0}};
      const accuracy = forecast.accuracy || {{months_scored: 0, total: {{}}, by_category: [], corrections_applied: false}};
      const committedItems = committed.items || [];
      // Rent lands as a single step of roughly a month's rent on its due day, which on a
      // cumulative chart dwarfs everything else and flattens the variable spend the chart
      // exists to show. The aggregate view therefore leaves it out. It stays in the
      // headline totals and in both tables: this is a presentation choice, not a change
      // to the forecast. Selecting rent explicitly still draws it.
      const sumSeries = rows => {{
        const first = rows.find(row => (row.cumulative || []).length);
        const days = first ? first.cumulative.map(point => point.day) : [];
        return days.map((day, i) => {{
          const point = {{day}};
          ['actual', 'committed', 'low', 'medium', 'high'].forEach(key => {{
            const values = rows
              .map(row => (row.cumulative || [])[i])
              .filter(entry => entry && entry[key] != null)
              .map(entry => entry[key]);
            point[key] = values.length
              ? Math.round(values.reduce((a, b) => a + b, 0) * 100) / 100
              : null;
          }});
          return point;
        }});
      }};
      const rentRow = categories.find(row => row.category === 'rent');
      const rentItem = (forecast.committed && forecast.committed.items || [])
        .find(item => item.category === 'rent');
      const selected = forecastCategory === 'total'
        ? {{...forecast.total, label: 'Total', cumulative: sumSeries(categories.filter(row => row.category !== 'rent'))}}
        : categories.find(row => row.category === forecastCategory);
      // Read the figure from the payload so it follows a rent change rather than drifting.
      const rentNote = forecastCategory === 'total' && rentRow
        ? `<div class="subtle">Excludes rent (${{money(rentRow.medium)}}${{rentItem && rentItem.expected_day ? `, due day ${{rentItem.expected_day}}` : ''}}). It is constant, and including it hides the shape of everything else.</div>`
        : '';
      if (!selected) {{ forecastCategory = 'total'; return renderForecast(); }}
      const options = `<option value="total">Total spend</option>` + categories.map(row =>
        `<option value="${{esc(row.category)}}" ${{row.category === forecastCategory ? 'selected' : ''}}>${{esc(row.label)}}${{row.net_negative ? ' — net of reimbursements' : ''}}</option>`
      ).join('');
      const profileCount = categories.filter(row => row.basis === 'profile').length;
      const chargedCount = committedItems.filter(item => item.status === 'charged').length;
      const forecastMonth = new Date(`${{forecast.month || '2000-01'}}-01T12:00:00`).toLocaleString(undefined, {{month: 'long'}});
      const seasonalInsight = row => {{
        const factor = Number(row.seasonal_factor || 1);
        if (factor >= 0.9 && factor <= 1.1) return '';
        const direction = factor > 1 ? 'higher' : 'lower';
        return `<div class="subtle">Typically ${{Math.round(Math.abs(factor - 1) * 100)}}% ${{direction}} in ${{esc(forecastMonth)}}</div>`;
      }};
      const accuracyPercent = value => value == null
        ? '—'
        : `${{Number(value) >= 0 ? '+' : ''}}${{(Number(value) * 100).toFixed(1)}}%`;
      const coveragePercent = value => value == null ? '—' : `${{(Number(value) * 100).toFixed(0)}}%`;
      const categoryAccuracy = row => {{
        const values = row.accuracy || {{months: 0}};
        if (!values.months) return '<span class="subtle">Not scored</span>';
        return `<span>${{accuracyPercent(values.median_pct_error)}} bias</span><div class="subtle">${{coveragePercent(values.in_band_rate)}} in band · ${{values.months}} mo</div>`;
      }};
      const lastAccuracy = accuracy.last_month;
      const learningNote = Number(accuracy.months_scored || 0) < 2
        ? `${{Number(accuracy.months_scored || 0)}} completed month${{Number(accuracy.months_scored || 0) === 1 ? '' : 's'}} scored. At least 2 are required before any correction is applied; all correction factors are 1.0.`
        : `${{Number(accuracy.months_scored || 0)}} completed months scored. Bounded bias and range-width corrections are active where a category has at least 2 months.`;
      app.innerHTML = `
        <section class="account-title">
          <div><h2>${{esc(forecast.month || 'Current month')}} Forecast</h2><div class="subtle">Spend to date plus fixed merchant commitments and variable spend historically still to come.</div></div>
          <label>Chart category <select id="forecastCategory">${{options}}</select></label>
          ${{rentNote}}
        </section>
        <section class="grid kpis">
          ${{kpi('Spent to date', money(forecast.total.spent), `Through day ${{forecast.day_of_month || 0}} of ${{forecast.days_in_month || 0}}`)}}
          ${{kpi('Low', money(forecast.total.low), '25th-percentile category outcomes')}}
          ${{kpi('Medium', money(forecast.total.medium), 'Median category outcomes')}}
          ${{kpi('High', money(forecast.total.high), '75th-percentile category outcomes')}}
          ${{kpi('Day-type coverage', `${{profileCount}} / ${{categories.length}}`, 'Categories with 3+ historical months')}}
        </section>
        <section class="grid kpis" style="margin-top:16px">
          ${{kpi('Committed remaining', money(committed.remaining), `${{committedItems.length}} recurring items · ${{chargedCount}} charged this month (${{money(committed.charged_so_far)}})`)}}
          ${{kpi('Variable remaining', money(forecast.total.variable_medium), `${{money(forecast.total.variable_low)}}–${{money(forecast.total.variable_high)}} low-high range`)}}
          ${{kpi('Weekdays remaining', Number(calendar.weekdays_remaining || 0).toLocaleString(), 'Calendar days after today')}}
          ${{kpi('Weekend days remaining', Number(calendar.weekend_days_remaining || 0).toLocaleString(), 'Saturdays and Sundays after today')}}
        </section>
        <section class="card" style="margin-top:16px">
          <h2>How this forecast has been doing</h2>
          <div class="subtle" style="margin-bottom:10px">${{esc(learningNote)}}</div>
          ${{lastAccuracy ? `<div class="grid kpis">
            ${{kpi('Last completed month', esc(lastAccuracy.month), `${{lastAccuracy.in_band ? 'Landed inside' : 'Landed outside'}} the predicted range`)}}
            ${{kpi('Predicted medium', money(lastAccuracy.predicted_medium), `Model ${{esc(accuracy.model_version || '')}}`)}}
            ${{kpi('Actual', money(lastAccuracy.actual), `${{accuracyPercent(lastAccuracy.pct_error)}} vs prediction`)}}
            ${{kpi('Range coverage', coveragePercent(accuracy.total && accuracy.total.in_band_rate), `${{accuracyPercent(accuracy.total && accuracy.total.median_pct_error)}} median error`)}}
          </div>` : '<div class="empty">No completed forecast month has been captured yet.</div>'}}
        </section>
        <section class="card" style="margin-top:16px">
          <h2>Committed Items</h2>
          <div class="subtle" style="margin-bottom:8px">Recurring merchant charges detected from the trailing six complete months. Overdue items remain in the forecast until they arrive.</div>
          ${{committedItems.length ? `<div class="table-wrap"><table><thead><tr><th>Merchant</th><th class="money">Expected amount</th><th class="money">Expected day</th><th>Status</th></tr></thead><tbody>
            ${{committedItems.map(item => `<tr class="${{item.status === 'overdue' ? 'forecast-overdue' : ''}}"><td>${{esc(item.merchant_label)}}<div class="subtle">${{esc(item.label)}}</div></td><td class="money">${{exactMoney(item.expected_amount)}}</td><td class="money">Day ${{item.expected_day}}</td><td><span class="tag forecast-status-${{esc(item.status)}}">${{esc(item.status)}}</span></td></tr>`).join('')}}
          </tbody></table></div>` : '<div class="empty">No committed items detected.</div>'}}
        </section>
        <section class="card" style="margin-top:16px">
          <div class="account-title"><div><h2>Cumulative ${{esc(selected.label)}} Spend</h2><div class="subtle">Actual through today; the dashed committed path shows spend already spoken for, while the variable range drives the forecast width.${{selected.net_negative ? ' Values are net of reimbursements.' : ''}}</div></div></div>
          ${{selected.cumulative?.length ? forecastChart(selected.cumulative) : '<div class="empty">No forecast series available.</div>'}}
        </section>
        <section class="card" style="margin-top:16px">
          <h2>Category Forecasts</h2>
          <div class="subtle" style="margin-bottom:8px">Day-type rows apply weekday and weekend daily-rate quartiles from the trailing six complete months. Run-rate rows have fewer than three complete months and do not imply the same precision.</div>
          <div class="table-wrap"><table><thead><tr><th>Category</th><th class="money">Spent</th><th class="money">Committed</th><th class="money">Variable median</th><th class="money">Low</th><th class="money">Medium</th><th class="money">High</th><th>Accuracy</th><th>Basis</th><th class="money">History</th></tr></thead><tbody>
            ${{categories.map(row => `<tr><td>${{esc(row.label)}}${{row.net_negative ? '<div class="subtle">Net of reimbursements</div>' : ''}}${{seasonalInsight(row)}}</td><td class="money">${{exactMoney(row.spent)}}</td><td class="money">${{exactMoney(row.committed)}}</td><td class="money">${{exactMoney(row.variable_medium)}}</td><td class="money">${{exactMoney(row.low)}}</td><td class="money">${{exactMoney(row.medium)}}</td><td class="money">${{exactMoney(row.high)}}</td><td>${{categoryAccuracy(row)}}</td><td><span class="tag">${{row.basis === 'profile' ? 'Day type' : 'Run rate'}}</span></td><td class="money">${{row.months_of_history}}</td></tr>`).join('')}}
          </tbody></table></div>
        </section>`;
      document.getElementById('forecastCategory').addEventListener('change', event => {{
        forecastCategory = event.target.value;
        renderForecast();
      }});
    }}

    function renderMonthly() {{
      const globalMonths = selectedGlobalMonths();
      const selectedMonths = selectedMonthlyMonths();
      const s = monthlySummary(selectedMonths);
      const unsupportedMonths = globalMonths.filter(month => !monthlySourceMonths().includes(month));
      const periodNote = unsupportedMonths.length
        ? `Monthly matrix uses workbook-summary months only; excluded from matrix: ${{unsupportedMonths.join(', ')}}.`
        : 'Monthly matrix is filtered to the global selected period.';
      app.innerHTML = `
        <section class="grid kpis">
          ${{kpi('Selected months', selectedMonths.length.toLocaleString(), selectedPeriodLabel(selectedMonths))}}
          ${{kpi('Salary', money(s.income), 'Selected range')}}
          ${{kpi('Expenses', money(s.expenses), `Avg ${{money(s.expenses / Math.max(selectedMonths.length, 1))}} per month`)}}
          ${{kpi('Investments', money(s.savings), 'Selected range')}}
          ${{kpi('Net Income', money(s.netIncome), 'Before investment savings')}}
          ${{kpi('Total Savings', money(s.totalSavings), 'Net income plus investments')}}
          ${{kpi('Transactions', s.transactions.toLocaleString(), 'Selected range')}}
        </section>
        <section class="card" style="margin-top:16px">
          <div class="account-title">
            <div><h2>Monthly Category Matrix</h2><div class="subtle">${{esc(periodNote)}}</div></div>
          </div>
          ${{selectedMonths.length ? monthlyMatrixTable() : '<div class="empty">No workbook-summary months match the selected period.</div>'}}
        </section>
        <section class="grid split" id="monthlySplitGrid" style="margin-top:16px;grid-template-columns:${{(monthlySplit * 100).toFixed(2)}}% 10px minmax(0, 1fr)">
          <div class="card"><h2>Range Category Rollup</h2><div class="subtle" style="margin-bottom:8px">Click a category to see the transactions behind it.</div>${{rollupBars(workbookMonthlyRows(selectedMonths).filter(r => r.style === 'normal' && r.kind === 'money' && r.label && !['Rent','Salary','401k','IRA','Additional Savings'].includes(r.label)), selectedMonths)}}</div>
          <div class="split-handle" id="monthlySplitHandle" role="separator" aria-orientation="vertical" aria-label="Resize the rollup and transaction panels" tabindex="0"></div>
          <div class="card"><h2>Range Transactions</h2>${{txTable(filtered().slice(0, 150))}}</div>
        </section>`;

      wireSplitHandle();
      app.querySelectorAll('[data-rollup-cat]').forEach(el => {{
        const toggle = () => {{
          const cat = el.dataset.rollupCat;
          monthlyRollupCat = monthlyRollupCat === cat ? null : cat;
          renderMonthly();
        }};
        el.addEventListener('click', toggle);
        el.addEventListener('keydown', event => {{
          if (event.key === 'Enter' || event.key === ' ') {{
            event.preventDefault();
            toggle();
          }}
        }});
      }});
    }}

    function renderAnalytics() {{
      const rows = filtered();
      const categoryRows = categoryRollup(rows);

      // Section 1: Month-over-month deltas
      const months = allTransactionMonths();
      const selMonths = selectedGlobalMonths();
      const cur = selMonths.at(-1);
      const curIdx = months.indexOf(cur);
      const prev = curIdx > 0 ? months[curIdx - 1] : null;
      let deltaHtml;
      if (selMonths.length < 2) {{
        deltaHtml = '<div class="empty">Need at least two months in the selected range.</div>';
      }} else {{
        const monthCat = {{}};
        const monthTotal = {{}};
        const monthInc = {{}};
        DATA.transactions.forEach(tx => {{
          const ec = expenseCost(tx);
          if (ec) {{
            (monthCat[tx.month] = monthCat[tx.month] || {{}});
            monthCat[tx.month][tx.category] = (monthCat[tx.month][tx.category] || 0) + ec;
            monthTotal[tx.month] = (monthTotal[tx.month] || 0) + ec;
          }}
          const iv = incomeValue(tx);
          if (iv) monthInc[tx.month] = (monthInc[tx.month] || 0) + iv;
        }});
        const curCats = monthCat[cur] || {{}};
        const prevCats = (prev && monthCat[prev]) || {{}};
        const topCats = Object.keys(curCats).sort((a, b) => curCats[b] - curCats[a]).slice(0, 8);
        const mk = (label, curV, prevV) => {{
          const isNew = prevV === 0;
          const pctChg = prevV > 0 ? (curV - prevV) / prevV : null;
          return deltaCard(label, curV, curV - prevV, pctChg, isNew);
        }};
        const deltaCards = topCats.map(cat => mk(categoryLabels[cat] || cat, curCats[cat] || 0, prevCats[cat] || 0));
        deltaCards.push(mk('Total Spend', monthTotal[cur] || 0, (prev && monthTotal[prev]) || 0));
        deltaCards.push(mk('Income', monthInc[cur] || 0, (prev && monthInc[prev]) || 0));
        deltaHtml = `<section class="grid kpis">${{deltaCards.join('')}}</section>`;
      }}

      // Section 2: Spend-over-time trend (search-filtered, all months)
      const trendQuery = searchInput.value.trim().toLowerCase();
      const matchesSearch = tx => !trendQuery || [tx.account, tx.date, tx.description, tx.category_label, tx.native_category, tx.counterparty].join(' ').toLowerCase().includes(trendQuery);
      const trendMap = {{}};
      months.forEach(m => trendMap[m] = {{month: m, expenses: 0, income: 0}});
      DATA.transactions.forEach(tx => {{
        if (!matchesSearch(tx)) return;
        const row = trendMap[tx.month];
        if (!row) return;
        row.expenses += expenseCost(tx);
        row.income += incomeValue(tx);
      }});
      const trendRows = months.map(m => trendMap[m]);
      const trendHtml = trendRows.length
        ? lineChart(trendRows, [{{key: 'expenses', color: '#2e6b3f', label: 'Spend'}}, {{key: 'income', color: '#9e7132', label: 'Income'}}])
        : '<div class="empty">No months to chart.</div>';

      // Section 3: Controllability lens for the visible period
      const controlOrder = ['high', 'medium', 'low'];
      const spendRows = rows.filter(tx => grossExpenseCost(tx) > 0);
      const controlRows = controlOrder.map(level => {{
        const matching = spendRows.filter(tx => controlLevel(tx.category) === level);
        const categories = new Set(matching.map(tx => tx.category));
        return {{
          level,
          label: controlLabels[level],
          expense: matching.reduce((sum, tx) => sum + grossExpenseCost(tx), 0),
          count: matching.length,
          categories: categories.size
        }};
      }});
      const controlTotal = controlRows.reduce((sum, row) => sum + row.expense, 0) || 1;
      const controlHtml = spendRows.length
        ? `<section class="grid three">${{controlRows.map(row => {{
            const share = row.expense / controlTotal;
            return `<div class="card control-card">
              <div class="chip-row"><span class="tag control-${{row.level}}">${{esc(row.label)}}</span></div>
              <div class="metric-value">${{money(row.expense)}}</div>
              <div class="control-bar"><div class="control-fill ${{row.level}}" style="width:${{Math.max(2, share * 100)}}%"></div></div>
              <div class="metric-note">${{pct(share)}} of visible spend · ${{row.categories}} categories · ${{controlNote(row.level)}}</div>
            </div>`;
          }}).join('')}}</section>`
        : '<div class="empty">No controllable spend matches the current filters.</div>';

      // Section 4: Anomaly / big-change flags
      const anomalyUniverse = selMonths.length ? selMonths : months;
      const latestDataMonth = months.at(-1);
      const completeMonths = anomalyUniverse.at(-1) === latestDataMonth ? anomalyUniverse.slice(0, -1) : anomalyUniverse;
      let anomalyHtml;
      if (completeMonths.length < 4) {{
        anomalyHtml = '<div class="empty">Not enough history for anomaly detection.</div>';
      }} else {{
        const completeSet = new Set(completeMonths);
        const catMonth = {{}};
        DATA.transactions.forEach(tx => {{
          if (!matchesSearch(tx)) return;
          const ec = expenseCost(tx);
          if (!ec || !completeSet.has(tx.month)) return;
          (catMonth[tx.category] = catMonth[tx.category] || {{}});
          catMonth[tx.category][tx.month] = (catMonth[tx.category][tx.month] || 0) + ec;
        }});
        const flags = [];
        Object.keys(catMonth).forEach(cat => {{
          const control = controlLevel(cat);
          if (control === 'internal') return;
          const series = completeMonths.map(m => catMonth[cat][m] || 0);
          const latest = series.at(-1);
          const prior = series.slice(-4, -1);
          const base = prior.reduce((a, b) => a + b, 0) / prior.length;
          const sd = Math.sqrt(prior.reduce((a, b) => a + (b - base) ** 2, 0) / prior.length);
          if (base >= 50 && (latest > base * 1.5 || latest > base + 2 * sd)) {{
            const high = latest >= base * 2 || latest > base + 3 * sd;
            flags.push({{cat, latest, base, high, control, delta: latest - base}});
          }}
        }});
        const controlRank = {{high: 0, medium: 1, low: 2}};
        const filteredFlags = flags
          .filter(f => analyticsControlFilter === 'all' || f.control === analyticsControlFilter)
          .sort((a, b) => (controlRank[a.control] - controlRank[b.control]) || ((b.latest - b.base) - (a.latest - a.base)));
        anomalyHtml = filteredFlags.length
          ? `<section class="insights">${{filteredFlags.map(f => {{
              const severity = f.high ? '<span class="sev high">High</span>' : '<span class="sev">Elevated</span>';
              const controlChip = `<span class="tag control-${{f.control}}">${{esc(controlLabels[f.control])}}</span>`;
              const above = f.base > 0 ? ` · ${{pct((f.latest - f.base) / f.base)}} above average` : '';
              return `<div class="card insight warning">
                <div class="chip-row">${{severity}}${{controlChip}}</div>
                <h3>${{esc(categoryLabels[f.cat] || f.cat)}}</h3>
                <p>${{money(f.latest)}} vs ${{money(f.base)}} trailing average${{above}}</p>
                <div class="metric-note">${{esc(controlRecommendation(f))}}</div>
              </div>`;
            }}).join('')}}</section>`
          : '<div class="empty">No categories in this control level deviated sharply for the selected period.</div>';
      }}

      // Section 5: Category deep-dive controls
      const deepDiveOptions = `<option value="" ${{analyticsCategory ? '' : 'selected'}}>Choose a category</option>` +
        categoryChoices.map(c => `<option value="${{c}}" ${{c === analyticsCategory ? 'selected' : ''}}>${{esc(categoryLabels[c])}}</option>`).join('');
      const controlOptions = [
        ['all', 'All control levels'],
        ['high', 'High control'],
        ['medium', 'Medium control'],
        ['low', 'Low control']
      ].map(([value, label]) => `<option value="${{value}}" ${{analyticsControlFilter === value ? 'selected' : ''}}>${{label}}</option>`).join('');

      app.innerHTML = `
        <section class="card">
          <h2>Spend Over Time</h2>
          <div class="subtle" style="margin-bottom:8px">Full history, current search applied.</div>
          ${{trendHtml}}
        </section>
        <section style="margin-top:16px">
          <h2>Controllability Lens</h2>
          <div class="subtle" style="margin-bottom:12px">Visible spend grouped by how directly you can change it.</div>
          ${{controlHtml}}
        </section>
        <section style="margin-top:16px">
          <h2>Month Over Month</h2>
          <div class="subtle" style="margin-bottom:12px">${{prev ? `${{cur}} versus ${{prev}}` : selectedPeriodLabel()}}</div>
          ${{deltaHtml}}
        </section>
        <section class="card" style="margin-top:16px">
          <h2>Category Deep Dive</h2>
          <div class="subtle" style="margin-bottom:8px">Pick a category to see its trend, share, and top merchants.</div>
          <select id="deepDiveCat" aria-label="Category to inspect">${{deepDiveOptions}}</select>
          <div id="deepDivePanel" style="margin-top:16px"></div>
        </section>
        <section style="margin-top:16px">
          <div class="account-title">
            <div><h2>Big Changes</h2><div class="subtle">Categories above their recent trailing average, prioritized by controllability.</div></div>
            <select id="controlFilter" aria-label="Filter big changes by controllability">${{controlOptions}}</select>
          </div>
          ${{anomalyHtml}}
        </section>
        <section class="grid two" style="margin-top:16px">
          <div class="card"><h2>Expense Concentration</h2>${{bars(categoryRows, 'expense', 14)}}</div>
          <div class="card"><h2>All Matching Transactions</h2>${{txTable(rows.slice(0, 200))}}</div>
        </section>`;

      const deepDiveSelect = document.getElementById('deepDiveCat');
      if (deepDiveSelect) {{
        deepDiveSelect.addEventListener('change', event => {{
          analyticsCategory = event.target.value || null;
          renderDeepDive();
        }});
      }}
      const controlFilter = document.getElementById('controlFilter');
      if (controlFilter) {{
        controlFilter.addEventListener('change', event => {{
          analyticsControlFilter = event.target.value;
          renderAnalytics();
        }});
      }}
      renderDeepDive();
    }}

    function renderDeepDive() {{
      const panel = document.getElementById('deepDivePanel');
      if (!panel) return;
      if (!analyticsCategory) {{
        panel.innerHTML = '<div class="empty">Select a category to inspect its trend and merchants.</div>';
        return;
      }}
      const cat = analyticsCategory;
      const perMonthRows = allTransactionMonths().map(m => {{
        let exp = 0;
        DATA.transactions.forEach(tx => {{ if (tx.category === cat && tx.month === m) exp += expenseCost(tx); }});
        return {{month: m, expenses: exp}};
      }});
      const allRows = filtered();
      const catRows = allRows.filter(tx => tx.category === cat);
      const catTotal = catRows.reduce((s, tx) => s + expenseCost(tx), 0);
      const totalExpense = allRows.reduce((s, tx) => s + expenseCost(tx), 0);
      const merchantMap = {{}};
      catRows.forEach(tx => {{
        const key = merchantGroupLabel(tx);
        if (!merchantMap[key]) merchantMap[key] = {{label: key, expense: 0, count: 0, variants: new Set()}};
        merchantMap[key].expense += expenseCost(tx);
        merchantMap[key].count += expenseCost(tx) > 0 ? 1 : 0;
        merchantMap[key].variants.add(tx.counterparty || tx.description);
      }});
      const merchantRows = Object.values(merchantMap)
        .map(row => ({{...row, variantCount: row.variants.size}}))
        .sort((a, b) => b.expense - a.expense);
      const merchants = merchantRows.some(r => r.expense > 0)
        ? `<div class="scroll-panel">${{bars(merchantRows, 'expense', merchantRows.length, null)}}</div>`
        : '<div class="empty">No merchant spend for this category in the current filters.</div>';
      panel.innerHTML = `
        ${{lineChart(perMonthRows)}}
        <section class="grid kpis" style="margin-top:16px">
          ${{kpi('Share of spend', totalExpense > 0 ? pct(catTotal / totalExpense) : '0%', 'Of visible spend')}}
          ${{kpi('Merchant groups', merchantRows.filter(r => r.expense > 0).length.toLocaleString(), 'Normalized labels')}}
        </section>
        <div style="margin-top:16px">${{merchants}}</div>`;
    }}

    function renderReview() {{
      const rows = uncertainTransactions();
      const overrideCount = Object.keys(categoryOverrides).length;
      app.innerHTML = `
        <section class="grid kpis">
          ${{kpi('Needs review', rows.length.toLocaleString(), selectedPeriodLabel())}}
          ${{kpi('Saved overrides', overrideCount.toLocaleString(), overrideStorageNote())}}
          ${{kpi('Uncategorized', rows.filter(tx => tx.category === 'uncategorized').length.toLocaleString(), 'No rule matched')}}
          ${{kpi('Food or bar?', rows.filter(tx => Number(tx.needs_review) === 1 && tx.category !== 'uncategorized').length.toLocaleString(), 'Venue could be either')}}
        </section>
        <section class="card" style="margin-top:16px">
          <div class="account-title">
            <div><h2>Category Review</h2><div class="subtle">Pick a category for uncertain rows. Rows drop off this list once you set one. ${{overrideBackend === 'server' ? 'Saved to <strong>financials.sqlite</strong>.' : 'Remembered in this browser only.'}}</div></div>
          </div>
          ${{reviewTable(rows)}}
        </section>`;
      wireCategorySelects();
    }}

    function renderAccount(account) {{
      const rows = filtered(account);
      const expense = rows.reduce((sum, tx) => sum + grossExpenseCost(tx), 0);
      const spendRowCount = rows.reduce((count, tx) => count + (expenseCost(tx) !== 0 ? 1 : 0), 0);
      const transfers = rows.reduce((sum, tx) => sum + transferCost(tx), 0);
      const netActivity = rows.reduce((sum, tx) => sum + Number(tx.cost || 0), 0);
      const income = rows.reduce((sum, tx) => sum + (['salary'].includes(tx.category) ? Math.max(-Number(tx.cost || 0), 0) : 0), 0);
      const categoryRows = Object.values(rows.reduce((acc, tx) => {{
        if (!acc[tx.category]) acc[tx.category] = {{category: tx.category, label: tx.category_label, expense: 0, count: 0}};
        acc[tx.category].expense += grossExpenseCost(tx);
        acc[tx.category].count += 1;
        return acc;
      }}, {{}})).sort((a,b) => b.expense - a.expense);
      app.innerHTML = `
        <section class="account-title">
          <div><h2>${{esc(account)}}</h2><div class="subtle">${{rows.length.toLocaleString()}} matching transactions</div></div>
        </section>
        <section class="grid kpis">
          ${{kpi('Transactions', rows.length.toLocaleString(), 'Current filters')}}
          ${{kpi('Gross Spend', money(expense), 'Excludes payments/transfers')}}
          ${{kpi('Payments/Transfers', money(transfers), 'Internal account movement')}}
          ${{kpi('Net Activity', money(netActivity), 'All visible ledger costs')}}
          ${{kpi('Income', money(income), 'Income categories only')}}
          ${{kpi('Avg Gross Spend', money(expense / Math.max(spendRowCount, 1)), 'Per spend row')}}
        </section>
        <section class="grid two" style="margin-top:16px">
          <div class="card"><h2>Transactions</h2><div class="subtle" style="margin-bottom:8px">Change a category inline; it applies across every tab. ${{overrideBackend === 'server' ? 'Saved to <strong>financials.sqlite</strong>, so it survives rebuilds.' : 'Remembered in this browser only &mdash; run <code>python3 serve_dashboard.py</code> to save edits to the database instead.'}}</div>${{txTable(rows, true)}}</div>
          <div class="card"><h2>Category Mix</h2>${{bars(categoryRows, 'expense', 14)}}</div>
        </section>`;
      wireCategorySelects();
    }}

    // Waterfall: opening balance, each signed flow stacked on the running total, closing
    // balance. Totals are drawn from the baseline, steps float at the running height.
    function waterfall(opening, steps, closing) {{
      const w = 920, h = 300, padL = 64, padR = 24, padTop = 20, padBot = 62;
      const items = [{{label: 'Opening', value: opening, total: true}}]
        .concat(steps.filter(s => Math.abs(s.value) >= 0.005))
        .concat([{{label: 'Closing', value: closing, total: true}}]);
      let run = 0;
      const spans = items.map(it => {{
        if (it.total) {{ run = it.value; return {{...it, lo: 0, hi: it.value}}; }}
        const lo = run, hi = run + it.value;
        run = hi;
        return {{...it, lo: Math.min(lo, hi), hi: Math.max(lo, hi)}};
      }});
      const maxV = Math.max(...spans.map(s => s.hi), 0);
      const minV = Math.min(...spans.map(s => s.lo), 0);
      const span = (maxV - minV) || 1;
      const yAt = v => padTop + (maxV - v) / span * (h - padTop - padBot);
      const bw = (w - padL - padR) / items.length * 0.62;
      const xAt = i => padL + (i + 0.5) * (w - padL - padR) / items.length;
      const zero = yAt(0);
      const bars = spans.map((s, i) => {{
        const x = xAt(i) - bw / 2;
        const yTop = yAt(s.hi), yBot = yAt(s.lo);
        const fill = s.total ? '#2f3b35' : (s.value >= 0 ? '#2e6b3f' : '#b13d3d');
        const tip = `${{tipText(s.label)}}<br>${{s.total ? '' : (s.value >= 0 ? '+' : '-')}}${{money(Math.abs(s.value))}}`
          + (s.total ? '' : `<br>running: ${{money(s.value >= 0 ? s.hi : s.lo)}}`);
        return `<rect x="${{x.toFixed(1)}}" y="${{yTop.toFixed(1)}}" width="${{bw.toFixed(1)}}" height="${{Math.max(2, yBot - yTop).toFixed(1)}}" fill="${{fill}}" rx="2" data-tip="${{tip}}"></rect>`;
      }}).join('');
      const labels = spans.map((s, i) => {{
        const short = s.label.length > 11 ? s.label.slice(0, 10) + '\\u2026' : s.label;
        return `<text x="${{xAt(i).toFixed(1)}}" y="${{h - 38}}" text-anchor="middle" font-size="10" fill="#64706b">${{esc(short)}}</text>`
          + `<text x="${{xAt(i).toFixed(1)}}" y="${{h - 24}}" text-anchor="middle" font-size="10" fill="${{s.value < 0 && !s.total ? '#b13d3d' : '#2f3b35'}}">${{moneyShort(s.value)}}</text>`;
      }}).join('');
      return `<svg class="chart" viewBox="0 0 ${{w}} ${{h}}" role="img" aria-label="Cash flow waterfall">
        <line x1="${{padL}}" y1="${{zero.toFixed(1)}}" x2="${{w - padR}}" y2="${{zero.toFixed(1)}}" stroke="var(--line)" stroke-width="1"></line>
        ${{bars}}${{labels}}
      </svg>`;
    }}

    function renderCashflow() {{
      const cf = DATA.cashflow || {{months: [], anchor: {{}}}};
      const sel = selectedGlobalMonths(cf.months.map(r => r.month));
      const rows = cf.months.filter(r => sel.includes(r.month));
      if (!rows.length) {{ app.innerHTML = '<div class="empty">No months in the selected period.</div>'; return; }}
      const cur = rows.at(-1);
      const anchor = cf.anchor || {{}};
      const signed = v => (v >= 0 ? '+' : '-') + money(Math.abs(v));
      const gap = Math.round((cur.net_income - cur.delta) * 100) / 100;

      const balanceRows = rows.filter(r => r.closing != null);
      const flowTable = `<table><thead><tr><th>Month</th><th class="money">Opening</th><th class="money">Salary</th>
          <th class="money">Transfers in</th><th class="money">Card payments</th><th class="money">Rent</th>
          <th class="money">Investments</th><th class="money">Other</th><th class="money">Change</th><th class="money">Closing</th></tr></thead><tbody>
        ${{[...rows].reverse().map(r => `<tr>
          <td>${{esc(r.month)}}</td>
          <td class="money">${{r.opening == null ? '&mdash;' : exactMoney(r.opening)}}</td>
          <td class="money">${{exactMoney(r.salary_in)}}</td>
          <td class="money ${{r.transfers_in < 0 ? 'negative' : ''}}">${{exactMoney(r.transfers_in)}}</td>
          <td class="money negative">${{exactMoney(r.card_payments)}}</td>
          <td class="money negative">${{exactMoney(r.rent)}}</td>
          <td class="money negative">${{exactMoney(r.investments)}}</td>
          <td class="money ${{r.other_cash < 0 ? 'negative' : ''}}">${{exactMoney(r.other_cash)}}</td>
          <td class="money ${{r.delta < 0 ? 'negative' : ''}}"><strong>${{exactMoney(r.delta)}}</strong></td>
          <td class="money">${{r.closing == null ? '&mdash;' : exactMoney(r.closing)}}</td>
        </tr>`).join('')}}</tbody></table>`;

      const bridgeRow = (label, value, note) => `<tr><td>${{esc(label)}}</td>
        <td class="money ${{value < 0 ? 'negative' : ''}}">${{exactMoney(value)}}</td><td class="subtle">${{esc(note)}}</td></tr>`;

      app.innerHTML = `
        <section class="grid kpis">
          ${{kpi('Checking balance', cur.closing == null ? '&mdash;' : money(cur.closing), `End of ${{cur.month}}`)}}
          ${{kpi('Change this month', signed(cur.delta), cur.delta < 0 ? 'Balance fell' : 'Balance rose')}}
          ${{kpi('Net income (P&L)', money(cur.net_income), 'Income minus spend, when charged')}}
          ${{kpi('P&L vs cash gap', signed(gap), 'Explained in the bridge below')}}
          ${{kpi('Card float', money(cur.float_cumulative), 'Charged, not yet paid (since start)')}}
        </section>

        <section class="card" style="margin-top:16px">
          <h2>Checking Balance</h2>
          <div class="subtle" style="margin-bottom:8px">The stock. Derived from one confirmed balance of
            ${{exactMoney(anchor.balance || 0)}} on ${{esc(anchor.as_of || '')}}, walked forward and back through every
            checking transaction. Months before the anchor come from workbook-era rows and are indicative only.</div>
          ${{balanceRows.length ? lineChart(balanceRows, [{{key: 'closing', color: '#2e6b3f', label: 'Balance'}}]) : '<div class="empty">No derived balances in range.</div>'}}
        </section>

        <section class="card" style="margin-top:16px">
          <h2>Where the money went &mdash; ${{esc(cur.month)}}</h2>
          <div class="subtle" style="margin-bottom:8px">The flows. Every movement in and out of checking, in order.</div>
          ${{waterfall(cur.opening == null ? 0 : cur.opening, [
            {{label: 'Salary', value: cur.salary_in}},
            {{label: 'Transfers in', value: cur.transfers_in}},
            {{label: 'Card payments', value: cur.card_payments}},
            {{label: 'Rent', value: cur.rent}},
            {{label: 'Investments', value: cur.investments}},
            {{label: 'Other', value: cur.other_cash}}
          ], cur.closing == null ? cur.delta : cur.closing)}}
        </section>

        <section class="grid two" style="margin-top:16px">
          <div class="card">
            <h2>P&amp;L to Cash Bridge &mdash; ${{esc(cur.month)}}</h2>
            <div class="subtle" style="margin-bottom:8px">Why net income and the balance change disagree.</div>
            <div class="table-wrap"><table><thead><tr><th>Line</th><th class="money">Amount</th><th>What it is</th></tr></thead><tbody>
              ${{bridgeRow('Net income', cur.net_income, 'Income minus spend, booked when charged')}}
              ${{bridgeRow('+ Spend not yet paid', cur.deferred_spend, 'Charged to cards or Venmo this month')}}
              ${{bridgeRow('- Card payments', cur.card_payments, "Cash out, settling earlier months' charges")}}
              ${{bridgeRow('- Investments', cur.investments, 'Moved to brokerage, not spent')}}
              ${{bridgeRow('+/- Transfers', cur.transfers_in, 'Venmo cashouts and account moves')}}
              ${{Math.abs(cur.bridge_residual) >= 0.02 ? bridgeRow('Unexplained', cur.bridge_residual, 'Does not reconcile - investigate') : ''}}
              <tr style="border-top:2px solid #1f2724"><td><strong>Checking change</strong></td>
                <td class="money ${{cur.delta < 0 ? 'negative' : ''}}"><strong>${{exactMoney(cur.delta)}}</strong></td>
                <td class="subtle">What the balance actually did</td></tr>
            </tbody></table></div>
          </div>
          <div class="card">
            <h2>Card Float</h2>
            <div class="subtle" style="margin-bottom:8px">Charges made but not yet paid off. A month with no card
              payment flatters the balance; the catch-up month punishes it.</div>
            <div class="table-wrap"><table><thead><tr><th>Month</th><th class="money">Charged</th><th class="money">Paid</th><th class="money">Float change</th></tr></thead><tbody>
              ${{[...rows].reverse().map(r => `<tr><td>${{esc(r.month)}}</td>
                <td class="money">${{exactMoney(r.charged_to_cards)}}</td>
                <td class="money">${{exactMoney(r.paid_to_cards)}}</td>
                <td class="money ${{r.float_change < 0 ? 'negative' : ''}}">${{exactMoney(r.float_change)}}</td></tr>`).join('')}}
            </tbody></table></div>
          </div>
        </section>

        <section class="card" style="margin-top:16px">
          <h2>Monthly Cash Flow</h2>
          <div class="subtle" style="margin-bottom:8px">Every checking flow by month. Opening plus the columns equals closing.</div>
          <div class="table-wrap">${{flowTable}}</div>
        </section>`;
    }}

    function render() {{
      [...tabs.children].forEach(btn => btn.classList.toggle('active', btn.dataset.tab === activeTab));
      if (activeTab === 'Cash Flow') return renderCashflow();
      if (activeTab === 'Overview') return renderOverview();
      if (activeTab === 'Forecast') return renderForecast();
      if (activeTab === 'Monthly') return renderMonthly();
      if (activeTab === 'Analytics') return renderAnalytics();
      if (activeTab === 'Review') return renderReview();
      return renderAccount(activeTab);
    }}

    function init() {{
      const chartTip = document.getElementById('chartTip');
      document.addEventListener('mousemove', e => {{
        const el = e.target.closest && e.target.closest('[data-tip]');
        if (el) {{
          chartTip.innerHTML = el.dataset.tip;
          const margin = 14;
          let x = e.clientX + margin;
          let y = e.clientY + margin;
          const rect = chartTip.getBoundingClientRect();
          if (x + rect.width > window.innerWidth - 8) x = e.clientX - margin - rect.width;
          if (y + rect.height > window.innerHeight - 8) y = e.clientY - margin - rect.height;
          chartTip.style.left = Math.max(8, x) + 'px';
          chartTip.style.top = Math.max(8, y) + 'px';
          chartTip.classList.add('show');
          chartTip.setAttribute('aria-hidden', 'false');
        }} else {{
          chartTip.classList.remove('show');
          chartTip.setAttribute('aria-hidden', 'true');
        }}
      }});
      document.getElementById('rangeLine').textContent = `${{DATA.summary.overall.first_date}} to ${{DATA.summary.overall.last_date}} · generated ${{DATA.generated_at}}`;
      loadCategoryOverrides();
      applyCategoryOverrides();
      // Render immediately from localStorage, then upgrade to the database-backed store
      // if serve_dashboard.py answers. Deliberately not awaited: the page must not block
      // on a probe that fails outright when opened as a file.
      initOverrideBackend();
      DATA.accounts.forEach(name => {{
        const btn = document.createElement('button');
        btn.textContent = name;
        btn.dataset.tab = name;
        btn.addEventListener('click', () => {{ activeTab = name; render(); }});
        tabs.appendChild(btn);
      }});
      const months = allTransactionMonths();
      monthSingle.innerHTML = monthOptionsFor(months, months.at(-1));
      monthStart.innerHTML = monthOptionsFor(months, months[0]);
      monthEnd.innerHTML = monthOptionsFor(months, months.at(-1));
      monthMode.value = 'all';
      updateMonthControlVisibility();
      searchInput.addEventListener('input', render);
      monthMode.addEventListener('change', () => {{ updateMonthControlVisibility(); render(); }});
      monthSingle.addEventListener('change', render);
      monthStart.addEventListener('change', render);
      monthEnd.addEventListener('change', render);
      document.getElementById('resetFilters').addEventListener('click', () => {{
        searchInput.value = '';
        monthMode.value = 'all';
        monthSingle.value = months.at(-1);
        monthStart.value = months[0];
        monthEnd.value = months.at(-1);
        updateMonthControlVisibility();
        render();
      }});
      render();
    }}
    init();
  </script>
</body>
</html>"""


def main() -> None:
    payload = build_payload()
    OUTPUT.write_text(render_html(payload), encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    print(f"{payload['summary']['overall']['transaction_count']} transactions across {payload['summary']['overall']['account_count']} accounts")


if __name__ == "__main__":
    main()
